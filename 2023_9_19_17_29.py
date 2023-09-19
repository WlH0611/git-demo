# coding: gbk
# coding: utf-8
#

from cgitb import text
from tkinter.simpledialog import askinteger
import uiautomator2 as u2
from uiautomator2 import Device
import imp
from sre_parse import State
import stat
import time
import tkinter
from tkinter import StringVar, ttk
from tkinter import *
import threading
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import cv2
import numpy as np
from PIL import Image
import requests
from io import BytesIO
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt

# 指定文件路径
file_path_Picture = "D:/Picture.xlsx"
file_path_Sound = "D:/Sound.xlsx"

try:
    # 尝试打开现有的工作簿
    workbook = load_workbook(file_path_Picture)
    sheet = workbook.active
except FileNotFoundError:
    # 如果文件不存在，则创建一个新的工作簿
    workbook = Workbook()
    sheet = workbook.active
    data = ["测试用例编号", "测试时间", "预期结果","测试结果","原因"]
    sheet.append(data)
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 80
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 80
    
try:
    # 尝试打开现有的工作簿
    workbook = load_workbook(file_path_Sound)
    sheet = workbook.active
except FileNotFoundError:
    # 如果文件不存在，则创建一个新的工作簿
    workbook = Workbook()
    sheet = workbook.active
    data = ["测试用例编号", "测试时间", "预期结果","测试结果","原因"]
    sheet.append(data)
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 80
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 80

if not os.path.exists("D:/auto_test_log"):
    # 创建文件夹
    os.makedirs("D:/auto_test_log")
    print("文件夹已创建")
else:
    print("文件夹已存在")

# 保存工作簿
workbook.save(file_path_Picture)
workbook.save(file_path_Sound)

d = u2.connect()

lock = threading.RLock()

top = tkinter.Tk()
w = top.winfo_screenwidth()
h = top.winfo_screenheight()
print(w) #1280
print(h) #720
top.state('zoomed')

frm = tkinter.Frame(top)
frm_left = tkinter.Frame(frm)
frm_right = tkinter.Frame(frm)

value = StringVar
auto_or_self = ttk.Combobox(frm_left,textvariable = value)
auto_or_self.configure(state = "readonly",width = 40)
auto_or_self.pack()
auto_or_self["value"] = ("自动测试全部","手动测试","自动测试Picture","自动测试Sound")
auto_or_self.current(0)

matrix = [1 for i in range(680)]
temp_list_Picture = []
temp_list_Picture.append("选择测试用例")
temp_list_Picture.append("Picture_PictureMode")
temp_list_Picture.append("into_Picture_Mode")
temp_list_Picture.append("Picture_ScreenBrightness")
temp_list_Picture.append("Picture_Brightness")
temp_list_Picture.append("Picture_Contrast")
temp_list_Picture.append("Picture_Saturation")
temp_list_Picture.append("Picture_HUE")
temp_list_Picture.append("Picture_Sharpness")
temp_list_Picture.append("Picture_AutoBrightness")
temp_list_Picture.append("Picture_Gamma")
temp_list_Picture.append("Picture_ColorTemperature_Mode")
temp_list_Picture.append("Picture_ColorTemperature_modify")
temp_list_Picture.append("Picture_ColorTemperature_modify_to_user")
temp_list_Picture.append("Picture_PWBC")
temp_list_Picture.append("Picture_AdvancedSettings")
temp_list_Picture.append("Picture_ColorTuner")
temp_list_Picture.append("Picture_Reset")
temp_list_Picture.append("Picture_Mode_compare_screen")
temp_list_Picture.append("Picture_Mode_cool_Temperature_compare_screen")



temp_list_All = []
temp_list_All.append("选择测试用例")
temp_list_All.append("Picture_PictureMode")
temp_list_All.append("into_Picture_Mode")
temp_list_All.append("Picture_ScreenBrightness")
temp_list_All.append("Picture_Brightness")
temp_list_All.append("Picture_Contrast")
temp_list_All.append("Picture_Saturation")
temp_list_All.append("Picture_HUE")
temp_list_All.append("Picture_Sharpness")
temp_list_All.append("Picture_AutoBrightness")
temp_list_All.append("Picture_Gamma")
temp_list_All.append("Picture_ColorTemperature_Mode")
temp_list_All.append("Picture_ColorTemperature_modify")
temp_list_All.append("Picture_ColorTemperature_modify_to_user")
temp_list_All.append("Picture_PWBC")
temp_list_All.append("Picture_AdvancedSettings")
temp_list_All.append("Picture_ColorTuner")
temp_list_All.append("Picture_Reset")
temp_list_All.append("Picture_Mode_compare_screen")
temp_list_All.append("Picture_Mode_cool_Temperature_compare_screen")
temp_list_All.append("Sound_Default")
temp_list_All.append("Sound_DolbyAtmos")
temp_list_All.append("Sound_SoundMode")
temp_list_All.append("Sound_DolbySoundMode")
temp_list_All.append("Sound_Balance_Bass_Treble")
temp_list_All.append("Sound_DTS")
temp_list_All.append("Sound_AutoVolume")
temp_list_All.append("Sound_Speaker")
temp_list_All.append("Sound_AudioType")
temp_list_All.append("Sound_VisuallyImpired")
temp_list_All.append("Sound_DialogueEnhancer")
temp_list_All.append("Sound_Reset")
temp_list_All.append("test")

temp_list_Sound = []
temp_list_Sound.append("选择测试用例")
temp_list_Sound.append("Sound_Default")
temp_list_Sound.append("Sound_DolbyAtmos")
temp_list_Sound.append("Sound_SoundMode")
temp_list_Sound.append("Sound_DolbySoundMode")
temp_list_Sound.append("Sound_Balance_Bass_Treble")
temp_list_Sound.append("Sound_DTS")
temp_list_Sound.append("Sound_AutoVolume")
temp_list_Sound.append("Sound_Speaker")
temp_list_Sound.append("Sound_AudioType")
temp_list_Sound.append("Sound_VisuallyImpired")
temp_list_Sound.append("Sound_DialogueEnhancer")
temp_list_Sound.append("Sound_Reset")

##模式转换需优化

choose_test = ttk.Combobox(frm_left)
choose_test.configure(width = 40)
choose_test.pack()
choose_test['value'] = temp_list_All
choose_test.state(['disabled'])
choose_test.current(0)





def aHash(img):
    # 均值哈希算法
    # 缩放为8*8
    img = cv2.resize(img, (8, 8))
    # 转换为灰度图
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # s为像素和初值为0，hash_str为hash值初值为''
    s = 0
    hash_str = ''
    # 遍历累加求像素和
    for i in range(8):
        for j in range(8):
            s = s+gray[i, j]
    # 求平均灰度
    avg = s/64
    # 灰度大于平均值为1相反为0生成图片的hash值
    for i in range(8):
        for j in range(8):
            if gray[i, j] > avg:
                hash_str = hash_str+'1'
            else:
                hash_str = hash_str+'0'
    return hash_str
 
def dHash(img):
    # 差值哈希算法
    # 缩放8*8
    img = cv2.resize(img, (9, 8))
    # 转换灰度图
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    hash_str = ''
    # 每行前一个像素大于后一个像素为1，相反为0，生成哈希
    for i in range(8):
        for j in range(8):
            if gray[i, j] > gray[i, j+1]:
                hash_str = hash_str+'1'
            else:
                hash_str = hash_str+'0'
    return hash_str
 
def pHash(img):
    # 感知哈希算法
    # 缩放32*32
    img = cv2.resize(img, (32, 32))   # , interpolation=cv2.INTER_CUBIC
 
    # 转换为灰度图
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # 将灰度图转为浮点型，再进行dct变换
    dct = cv2.dct(np.float32(gray))
    # opencv实现的掩码操作
    dct_roi = dct[0:8, 0:8]
 
    hash = []
    avreage = np.mean(dct_roi)
    for i in range(dct_roi.shape[0]):
        for j in range(dct_roi.shape[1]):
            if dct_roi[i, j] > avreage:
                hash.append(1)
            else:
                hash.append(0)
    return hash
 
def calculate(image1, image2):
    # 灰度直方图算法
    # 计算单通道的直方图的相似值
    hist1 = cv2.calcHist([image1], [0], None, [256], [0.0, 255.0])
    hist2 = cv2.calcHist([image2], [0], None, [256], [0.0, 255.0])
    # 计算直方图的重合度
    degree = 0
    for i in range(len(hist1)):
        if hist1[i] != hist2[i]:
            degree = degree + \
                (1 - abs(hist1[i] - hist2[i]) / max(hist1[i], hist2[i]))
        else:
            degree = degree + 1
    degree = degree / len(hist1)
    return degree
 
def classify_hist_with_split(image1, image2, size=(256, 256)):
    # RGB每个通道的直方图相似度
    # 将图像resize后，分离为RGB三个通道，再计算每个通道的相似值
    image1 = cv2.resize(image1, size)
    image2 = cv2.resize(image2, size)
    sub_image1 = cv2.split(image1)
    sub_image2 = cv2.split(image2)
    sub_data = 0
    for im1, im2 in zip(sub_image1, sub_image2):
        sub_data += calculate(im1, im2)
    sub_data = sub_data / 3
    return sub_data
 
def cmpHash(hash1, hash2):
    # Hash值对比
    # 算法中1和0顺序组合起来的即是图片的指纹hash。顺序不固定，但是比较的时候必须是相同的顺序。
    # 对比两幅图的指纹，计算汉明距离，即两个64位的hash值有多少是不一样的，不同的位数越小，图片越相似
    # 汉明距离：一组二进制数据变成另一组数据所需要的步骤，可以衡量两图的差异，汉明距离越小，则相似度越高。汉明距离为0，即两张图片完全一样
    n = 0
    # hash长度不同则返回-1代表传参出错
    if len(hash1) != len(hash2):
        return -1
    # 遍历判断
    for i in range(len(hash1)):
        # 不相等则n计数+1，n最终为相似度
        if hash1[i] != hash2[i]:
            n = n + 1
    return n
 
def getImageByUrl(url):
    # 根据图片url 获取图片对象
    html = requests.get(url, verify=False)
    image = Image.open(BytesIO(html.content))
    return image
 
def PILImageToCV():
    # PIL Image转换成OpenCV格式
    path = "/Users/waldenz/Documents/Work/doc/TestImages/t3.png"
    img = Image.open(path)
    plt.subplot(121)
    plt.imshow(img)
    print(isinstance(img, np.ndarray))
    img = cv2.cvtColor(np.asarray(img), cv2.COLOR_RGB2BGR)
    print(isinstance(img, np.ndarray))
    plt.subplot(122)
    plt.imshow(img)
    plt.show()
 
def CVImageToPIL():
    # OpenCV图片转换为PIL image
    path = "/Users/waldenz/Documents/Work/doc/TestImages/t3.png"
    img = cv2.imread(path)
    # cv2.imshow("OpenCV",img)
    plt.subplot(121)
    plt.imshow(img)
 
    img2 = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
    plt.subplot(122)
    plt.imshow(img2)
    plt.show()
 
def bytes_to_cvimage(filebytes):
    # 图片字节流转换为cv image
    image = Image.open(filebytes)
    img = cv2.cvtColor(np.asarray(image), cv2.COLOR_RGB2BGR)
    return img
 
def runAllImageSimilaryFun(para1, para2):
    # 均值、差值、感知哈希算法三种算法值越小，则越相似,相同图片值为0
    # 三直方图算法和单通道的直方图 0-1之间，值越大，越相似。 相同图片为1
 
    # t1,t2   14;19;10;  0.70;0.75
    # t1,t3   39 33 18   0.58 0.49
    # s1,s2  7 23 11     0.83 0.86  挺相似的图片
    # c1,c2  11 29 17    0.30 0.31
 
    if para1.startswith("http"):
         # 根据链接下载图片，并转换为opencv格式
        img1 = getImageByUrl(para1)
        img1 = cv2.cvtColor(np.asarray(img1), cv2.COLOR_RGB2BGR)
 
        img2 = getImageByUrl(para2)
        img2 = cv2.cvtColor(np.asarray(img2), cv2.COLOR_RGB2BGR)
    else:
        # 通过imread方法直接读取物理路径
        img1 = cv2.imread(para1)
        img2 = cv2.imread(para2)
 
    hash1 = aHash(img1)
    hash2 = aHash(img2)
    n1 = cmpHash(hash1, hash2)
    print('均值哈希算法相似度aHash：', n1)
 
    hash1 = dHash(img1)
    hash2 = dHash(img2)
    n2 = cmpHash(hash1, hash2)
    print('差值哈希算法相似度dHash：', n2)
 
    hash1 = pHash(img1)
    hash2 = pHash(img2)
    n3 = cmpHash(hash1, hash2)
    print('感知哈希算法相似度pHash：', n3)
 
    n4 = classify_hist_with_split(img1, img2)
    print('三直方图算法相似度：', n4)
 
    n5 = calculate(img1, img2)
    print("单通道的直方图", n5)
    #print("%d %d %d %.2f %.2f " % (n1, n2, n3, round(n4[0], 2), n5[0]))
    #print("%.2f %.2f %.2f %.2f %.2f " % (1-float(n1/64), 1 - float(n2/64), 1-float(n3/64), round(n4[0], 2), n5[0]))
 
    return n1,n2,n3





##进入Picture
def into_Picture():
    print("into_Picture")
    d.press("home")
    time.sleep(1)
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture")):
        d.press("down")
        time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture").click()
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    
##判断模式切换
def Picture_PictureMode():
    os.system("start d:\logcat.bat Picture_PictureMode")
    global sheet
    global workbook
    global file_path_Picture
    #Vivid
    into_Picture()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "Energy Saving":
        print("默认Energy Saving模式")
    else:
        print("未默认Energy Saving模式，错误")
        print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','默认非Energy Saving模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','默认非Energy Saving模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Vivid").click()
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "Vivid":
        print("切换Vivid模式成功")
    else:
        print("切换Vivid模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Vivid模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Vivid模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    #Stardand
    into_Picture()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "Vivid":
        print("记忆Vivid模式")
    else:
        print("未记忆Vivid模式，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Vivid模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Vivid模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Standard").click()
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "Standard":
        print("切换Standard模式成功")
    else:
        print("切换Standard模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Standard模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Standard模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    #Sport
    into_Picture()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "Standard":
        print("记忆Standard模式")
    else:
        print("未记忆Standard模式，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','未记忆Standard模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','未记忆Standard模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Sport").click()
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "Sport":
        print("切换Sport模式成功")
    else:
        print("切换Sport模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Sport模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Sport模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    #Movie
    into_Picture()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "Sport":
        print("记忆Sport模式")
    else:
        print("未记忆Sport模式，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','未记忆Sport模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','未记忆Sport模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Movie").click()
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "Movie":
        print("切换Movie模式成功")
    else:
        print("切换Movie模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Movie模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Movie模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    #User
    into_Picture()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "Movie":
        print("记忆Movie模式")
    else:
        print("未记忆Movie模式，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','未记忆Movie模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','未记忆Movie模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "User").click()
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "User":
        print("切换User模式成功")
    else:
        print("切换User模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换User模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换User模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    #Energy Saving
    into_Picture()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "User":
        print("记忆User模式")
    else:
        print("未记忆User模式，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','未记忆User模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','未记忆User模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Energy Saving").click()
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[2]').get_text() == "Energy Saving":
        print("切换Energy Saving模式成功")
    else:
        print("切换Energy Saving模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Energy Saving模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','错误','切换Energy Saving模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','正确','默认Energy Saving模式且切换后可记忆'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'默认Energy Saving模式且切换后可记忆','正确','默认Energy Saving模式且切换后可记忆'])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    d.press("back")
    time.sleep(1)
    try:
        os.remove('D:\\auto_test_log\Picture_PictureMode.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##判断模式切换时参数是否随之变化，且修改时是否将参数复制到User模式，且原模式参数不发生改变
def into_Picture_Mode():
    os.system("start d:\logcat.bat into_Picture_Mode")
    global sheet
    global workbook
    global file_path_Picture
    ##Vivid初始参数
    Vivid_ScreenBrightness = 100
    Vivid_Brightness = 50
    Vivid_Contrast = 60
    Vivid_Saturation = 60
    Vivid_HUE = 0
    Vivid_Sharpness = 12
    ##Standard初始化参数
    Standard_ScreenBrightness = 100
    Standard_Brightness = 50
    Standard_Contrast = 50
    Standard_Saturation = 50
    Standard_HUE = 0
    Standard_Sharpness = 10
    ##Sport初始参数
    Sport_ScreenBrightness = 100
    Sport_Brightness = 50
    Sport_Contrast = 55
    Sport_Saturation = 55
    Sport_HUE = 0
    Sport_Sharpness = 10
    ##Movie初始参数
    Movie_ScreenBrightness = 100
    Movie_Brightness = 50
    Movie_Contrast = 50
    Movie_Saturation = 40
    Movie_HUE = 0
    Movie_Sharpness = 8
    ##User初始参数
    User_ScreenBrightness = 100
    User_Brightness = 50
    User_Contrast = 50
    User_Saturation = 50
    User_HUE = 0
    User_Sharpness = 10
    ##EnergySaving初始参数
    EnergySaving_ScreenBrightness = 100
    EnergySaving_Brightness = 50
    EnergySaving_Contrast = 50
    EnergySaving_Saturation = 50
    EnergySaving_HUE = 0
    EnergySaving_Sharpness = 10
    
    into_Picture()
    
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "User").click()
    time.sleep(1)
    
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "User":
        print("Picture Mode进入User模式成功")
    else:
        print("Picture Mode进入User模式失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture Mode进入User模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture Mode进入User模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == User_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == User_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == User_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == User_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == User_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == User_Sharpness:
        print("Picture_User初始参数正确")
    else:
        print("Picture_User初始参数错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_User初始参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_User初始参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("up")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return

    ##----------------------Vivid-----------------------##
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Vivid").click()
    time.sleep(1)    
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_Sharpness:
        print("Picture_Vivid初始参数正确")
    else:
        print("Picture_Vivid初始参数错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Vivid初始参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Vivid初始参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("left")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("up")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "User":
        print("修改数据切换到User成功")
    else:
        print("修改数据切换到User失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','修改数据切换到User失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','修改数据切换到User失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Vivid").click()
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Vivid_Sharpness:
        print("Picture_Vivid参数修改到User，原数据不改变")
    else:
        print("Picture_Vivid参数修改到User，原数据改变")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Vivid参数修改到User，原数据改变'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Vivid参数修改到User，原数据改变'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("up")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return

    ##----------------------Standard-----------------------------##
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Standard").click()
    time.sleep(1)    
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_Sharpness:
        print("Picture_Standard初始参数正确")
    else:
        print("Picture_Standard初始参数错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Standard初始参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Standard初始参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("left")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("up")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "User":
        print("修改数据切换到User成功")
    else:
        print("修改数据切换到User失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','修改数据切换到User失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','修改数据切换到User失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Standard").click()
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Standard_Sharpness:
        print("Picture_Standard参数修改到User，原数据不改变")
    else:
        print("Picture_Standard参数修改到User，原数据改变")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Vivid参数修改到User，原数据改变'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Vivid参数修改到User，原数据改变'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("up")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return

    ##----------------------Sport------------------##
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Sport").click()
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_Sharpness:
        print("Picture_Sport初始参数正确")
    else:
        print("Picture_Sport初始参数错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Sport初始参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Sport初始参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("left")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("up")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "User":
        print("修改数据切换到User成功")
    else:
        print("修改数据切换到User失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','修改数据切换到User失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','修改数据切换到User失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Sport").click()
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Sport_Sharpness:
        print("Picture_Sport参数修改到User，原数据不改变")
    else:
        print("Picture_Sport参数修改到User，原数据改变")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Vivid参数修改到User，原数据改变'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Vivid参数修改到User，原数据改变'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("up")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return

    ##-----------------------------Movie------------------##
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Movie").click()
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_Sharpness:
        print("Picture_Movie初始参数正确")
    else:
        print("Picture_Sport初始参数错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Movie初始参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Movie初始参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("left")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("up")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "User":
        print("修改数据切换到User成功")
    else:
        print("修改数据切换到User失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','修改数据切换到User失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','修改数据切换到User失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Movie").click()
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Movie_Sharpness:
        print("Picture_Movie参数修改到User，原数据不改变")
    else:
        print("Picture_Movie参数修改到User，原数据改变")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Vivid参数修改到User，原数据改变'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_Vivid参数修改到User，原数据改变'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("up")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return

    ##----------------------EnergySaving-------------------##
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Standard").click()
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_Sharpness:
        print("Picture_EnergySaving初始参数正确")
    else:
        print("Picture_EnergySaving初始参数错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_EnergySaving初始参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_EnergySaving初始参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("left")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("up")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "User":
        print("修改数据切换到User成功")
    else:
        print("修改数据切换到User失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','修改数据切换到User失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','修改数据切换到User失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").click()
    time.sleep(1)
    d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Energy Saving").click()
    time.sleep(1)
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Sharpness")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Sharpness选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Sharpness选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_ScreenBrightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_Brightness and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_Contrast and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_Saturation and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_HUE and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == EnergySaving_Sharpness:
        print("Picture_EnergySaving参数修改到User，原数据不改变")
    else:
        print("Picture_EnergySaving参数修改到User，原数据改变")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_EnergySaving参数修改到User，原数据改变'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','Picture_EnergySaving参数修改到User，原数据改变'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode")):
        d.press("down")
        time.sleep(1)
        if(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Picture Reset")):
            print("没有Picture Mode选项")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','错误','没有Picture Mode选项'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','正确','Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变','正确','Picture_Mode初始参数正确，且修改参数自动复制到User，原数据不变'])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\into_Picture_Mode.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##进入Vivid模式
def into_Picture_Mode_Vivid():
    into_Picture()
    d.press("center")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("back")
    
##进入Standard模式
def into_Picture_Mode_Standard():
    into_Picture()
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("back")
    
##进入Sport模式
def into_Picture_Mode_Sport():
    into_Picture()
    d.press("center")
    time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("back")
    
##进入Movie模式
def into_Picture_Mode_Movie():
    into_Picture()
    d.press("center")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("back")
    
##进入User模式
def into_Picture_Mode_User():
    into_Picture()
    d.press("center")
    time.sleep(1)
    for i in range(4):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("back")
    
##进入EnergySaving模式
def into_Picture_Mode_EnergySaving():
    into_Picture()
    d.press("center")
    time.sleep(1)
    for i in range(5):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("back")

##修改ScreenBrightness
def Picture_ScreenBrightness():
    os.system("start d:\logcat.bat Picture_ScreenBrightness")
    global sheet
    global workbook
    global file_path_Picture
    into_Picture_Mode_User()
    global test_num
    
    into_Picture()
    for i in range(2):
        d.press("down")
        time.sleep(1)
    ScreenBrightness = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text())
    j = 5
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == ScreenBrightness:
        print("修改ScreenBrightness正确")
    else:
        print("修改ScreenBrightness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改ScreenBrightness正确并记忆",'错误',"修改ScreenBrightness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改ScreenBrightness正确并记忆",'错误',"修改ScreenBrightness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(2):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == ScreenBrightness:
        print("记忆ScreenBrightness正确")
    else:
        print("记忆ScreenBrightness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),"修改ScreenBrightness正确并记忆",'错误',"记忆ScreenBrightness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),"修改ScreenBrightness正确并记忆",'错误',"记忆ScreenBrightness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == ScreenBrightness:
        print("修改ScreenBrightness正确")
    else:
        print("修改ScreenBrightness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),"修改ScreenBrightness正确并记忆",'错误',"修改ScreenBrightness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),"修改ScreenBrightness正确并记忆",'错误',"修改ScreenBrightness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(2):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == ScreenBrightness:
        print("记忆ScreenBrightness正确")
    else:
        print("记忆ScreenBrightness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),"修改ScreenBrightness正确并记忆",'错误',"记忆ScreenBrightness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),"修改ScreenBrightness正确并记忆",'错误',"记忆ScreenBrightness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),"修改ScreenBrightness正确并记忆",'正确',"修改ScreenBrightness正确并记忆"))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),"修改ScreenBrightness正确并记忆",'正确',"修改ScreenBrightness正确并记忆"])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    d.press("back")
    try:
        os.remove('D:\\auto_test_log\Picture_ScreenBrightness.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
        
##修改Brightness
def Picture_Brightness():
    os.system("start d:\logcat.bat Picture_Brightness")
    global sheet
    global workbook
    global file_path_Picture
    global test_num
    into_Picture_Mode_User()
    into_Picture()
    for i in range(3):
        d.press("down")
        time.sleep(1)
    Brightness = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text())
    j = 5
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == Brightness:
        print("修改Brightness正确")
    else:
        print("修改Brightness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"修改Brightness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"修改Brightness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(3):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == Brightness:
        print("记忆Brightness正确")
    else:
        print("记忆Brightness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"记忆Brightness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"记忆Brightness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j * 2):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Brightness + j:
        print("修改Brightness正确")
    else:
        print("修改Brightness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"修改Brightness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"修改Brightness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(3):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Brightness + j:
        print("记忆Brightness正确")
    else:
        print("记忆Brightness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"记忆Brightness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"记忆Brightness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Brightness:
        print("修改Brightness正确")
    else:
        print("修改Brightness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"修改Brightness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"修改Brightness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(3):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Brightness:
        print("记忆Brightness正确")
    else:
        print("记忆Brightness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"记忆Brightness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'错误',"记忆Brightness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'正确',"修改Brightness正确并记忆"))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Brightness正确并记忆",'正确',"修改Brightness正确并记忆"])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    d.press("back")
    try:
        os.remove('D:\\auto_test_log\Picture_Brightness.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##修改Contrast
def Picture_Contrast():
    os.system("start d:\logcat.bat Picture_Contrast")
    global sheet
    global workbook
    global file_path_Picture
    global test_num
    into_Picture_Mode_User()
    
    into_Picture()
    for i in range(4):
        d.press("down")
        time.sleep(1)
    Contrast = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text())
    j = 5
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == Contrast:
        print("修改Contrast正确")
    else:
        print("修改Contrast错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"修改Contrast错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"修改Contrast错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(4):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == Contrast:
        print("记忆Contrast正确")
    else:
        print("记忆Contrast错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"记忆Contrast错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"记忆Contrast错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j * 2):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Contrast + j:
        print("修改Contrast正确")
    else:
        print("修改Contrast错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"修改Contrast错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"修改Contrast错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(4):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Contrast + j:
        print("记忆Contrast正确")
    else:
        print("记忆Contrast错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"记忆Contrast错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"记忆Contrast错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Contrast:
        print("修改Contrast正确")
    else:
        print("修改Contrast错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"修改Contrast错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"修改Contrast错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(4):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Contrast:
        print("记忆Contrast正确")
    else:
        print("记忆Contrast错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"记忆Contrast错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'错误',"记忆Contrast错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'正确',"修改Contrast正确并记忆"))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Contrast正确并记忆",'正确',"修改Contrast正确并记忆"])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    d.press("back")
    try:
        os.remove('D:\\auto_test_log\Picture_Contrast.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
        
##修改Saturation
def Picture_Saturation():
    os.system("start d:\logcat.bat Picture_Saturation")
    global sheet
    global workbook
    global file_path_Picture
    global test_num
    into_Picture_Mode_User()
    
    into_Picture()
    for i in range(5):
        d.press("down")
        time.sleep(1)
    Saturation = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text())
    j = 5
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == Saturation:
        print("修改Saturation正确")
    else:
        print("修改Saturation错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"修改Saturation错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"修改Saturation错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(5):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == Saturation:
        print("记忆Saturation正确")
    else:
        print("记忆Saturation错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"记忆Saturation错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"记忆Saturation错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j * 2):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Saturation + j:
        print("修改Saturation正确")
    else:
        print("修改Saturation错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"修改Saturation错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"修改Saturation错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(5):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Saturation + j:
        print("记忆Saturation正确")
    else:
        print("记忆Saturation错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"记忆Saturation错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"记忆Saturation错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Saturation:
        print("修改Saturation正确")
    else:
        print("修改Saturation错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"修改Saturation错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"修改Saturation错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(5):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Saturation:
        print("记忆Saturation正确")
    else:
        print("记忆Saturation错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"记忆Saturation错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'错误',"记忆Saturation错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'正确',"修改Saturation正确并记忆"))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Saturation正确并记忆",'正确',"修改Saturation正确并记忆"])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    d.press("back")
    try:
        os.remove('D:\\auto_test_log\Picture_Saturation.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##修改HUE
def Picture_HUE():
    os.system("start d:\logcat.bat Picture_HUE")
    global sheet
    global workbook
    global file_path_Picture
    global test_num
    into_Picture_Mode_User()
    
    into_Picture()
    for i in range(6):
        d.press("down")
        time.sleep(1)
    HUE = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text())
    j = 5
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == HUE:
        print("修改HUE正确")
    else:
        print("修改HUE错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"修改HUE错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"修改HUE错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(6):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == HUE:
        print("记忆HUE正确")
    else:
        print("记忆HUE错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"记忆HUE错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"记忆HUE错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j * 2):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == HUE + j:
        print("修改HUE正确")
    else:
        print("修改HUE错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"修改HUE错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"修改HUE错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(6):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == HUE + j:
        print("记忆HUE正确")
    else:
        print("记忆HUE错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"记忆HUE错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"记忆HUE错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == HUE:
        print("修改HUE正确")
    else:
        print("修改HUE错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"修改HUE错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"修改HUE错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(6):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == HUE:
        print("记忆HUE正确")
    else:
        print("记忆HUE错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"记忆HUE错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'错误',"记忆HUE错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'正确',"修改HUE正确并记忆"))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改HUE正确并记忆",'正确',"修改HUE正确并记忆"])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    d.press("back")
    try:
        os.remove('D:\\auto_test_log\Picture_HUE.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##修改Sharpness
def Picture_Sharpness():
    os.system("start d:\logcat.bat Picture_Sharpness")
    global sheet
    global workbook
    global file_path_Picture
    global test_num
    into_Picture_Mode_User()
    
    into_Picture()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    Sharpness = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text())
    j = 5
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == Sharpness:
        print("修改Sharpness正确")
    else:
        print("修改Sharpness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"修改Sharpness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"修改Sharpness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) + j == Sharpness:
        print("记忆Sharpness正确")
    else:
        print("记忆Sharpness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"记忆Sharpness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"记忆Sharpness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j * 2):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Sharpness + j:
        print("修改Sharpness正确")
    else:
        print("修改Sharpness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"修改Sharpness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"修改Sharpness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Sharpness + j:
        print("记忆Sharpness正确")
    else:
        print("记忆Sharpness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"记忆Sharpness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"记忆Sharpness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(j):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Sharpness:
        print("修改Sharpness正确")
    else:
        print("修改Sharpness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"修改Sharpness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"修改Sharpness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == Sharpness:
        print("记忆Sharpness正确")
    else:
        print("记忆Sharpness错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"记忆Sharpness错误"))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'错误',"记忆Sharpness错误"])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'正确',"修改Sharpness正确并记忆"))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"修改Sharpness正确并记忆",'正确',"修改Sharpness正确并记忆"])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    d.press("back")
    try:
        os.remove('D:\\auto_test_log\Picture_Sharpness.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##AutoBrightness开关设置
def Picture_AutoBrightness():
    os.system("start d:\logcat.bat Picture_AutoBrightness")
    global sheet
    global workbook
    global file_path_Picture
    into_Picture_Mode_Vivid()
    into_Picture()
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == False:
        print("Auto Brightness默认关闭")
    else:
        print("Auto Brightness默认开启，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Auto Brightness默认关闭且切换后可记忆','错误','Auto Brightness默认开启'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Auto Brightness默认关闭且切换后可记忆','错误','Auto Brightness默认开启'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    into_Picture()
    d.press("down")
    time.sleep(1)
    d.press("center")
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == True:
        print("Auto Brightness开启且记忆")
    else:
        print("Auto Brightness未开启，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Auto Brightness默认关闭且切换后可记忆','错误','Auto Brightness未开启'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Auto Brightness默认关闭且切换后可记忆','错误','Auto Brightness未开启'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Auto Brightness默认关闭且切换后可记忆','正确','Auto Brightness默认关闭且切换后可记忆'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Auto Brightness默认关闭且切换后可记忆','正确','Auto Brightness默认关闭且切换后可记忆'])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    d.press("center")
    time.sleep(1)
    d.press("home")
    into_Picture_Mode_EnergySaving()
    try:
        os.remove('D:\\auto_test_log\Picture_AutoBrightness.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##Gamma模式切换
def Picture_Gamma():
    os.system("start d:\logcat.bat Picture_Gamma")
    global sheet
    global workbook
    global file_path_Picture
    into_Picture_Mode_User()
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Middle":
        print("Picture_Gamma默认Middle")
    else:
        print("Picture_Gamma默认非Middle，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma默认非Middle'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma默认非Middle'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Dark":
        print("Picture_Gamma切换Dark成功")
    else:
        print("Picture_Gamma切换Dark失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma切换Dark失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma切换Dark失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Dark":
        print("Picture_Gamma记忆Dark成功")
    else:
        print("Picture_Gamma记忆Dark失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma记忆Dark失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma记忆Dark失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Bright":
        print("Picture_Gamma切换Bright成功")
    else:
        print("Picture_Gamma切换Bright失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma切换Bright失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma切换Bright失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Bright":
        print("Picture_Gamma记忆Bright成功")
    else:
        print("Picture_Gamma记忆Bright失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma记忆Bright失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma记忆Bright失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Middle":
        print("Picture_Gamma切换Middle成功")
    else:
        print("Picture_Gamma切换Middle失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma切换Middle失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','错误','Picture_Gamma切换Middle失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    os.system("adb shell kill `pgrep logcat`")
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','正确','Picture_Gamma默认Middle且切换后可记忆'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Gamma默认Middle且切换后可记忆','正确','Picture_Gamma默认Middle且切换后可记忆'])
    workbook.save(file_path_Picture)
    d.press("back")
    time.sleep(1)
    into_Picture_Mode_EnergySaving()
    try:
        os.remove('D:\\auto_test_log\Picture_Gamma.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##Color Temperature模式切换
def Picture_ColorTemperature_Mode():
    os.system("start d:\logcat.bat Picture_ColorTemperature")
    global sheet
    global workbook
    global file_path_Picture
    #User
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "Standard":
        print("Picture_ColorTemperature默认Standard模式")
    else:
        print("Picture_ColorTemperature默认非Standard模式，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature默认非Standard模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature默认非Standard模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "User":
        print("Picture_ColorTemperature切换User模式成功")
    else:
        print("Picture_ColorTemperature切换User模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature切换User模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature切换User模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    #Cool
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "User":
        print("Picture_ColorTemperature记忆User模式")
    else:
        print("Picture_ColorTemperature未记忆User模式，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature未记忆User模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature未记忆User模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "Cool":
        print("Picture_ColorTemperature切换Cool模式成功")
    else:
        print("Picture_ColorTemperature切换Cool模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature切换Cool模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature切换Cool模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    #Warm
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "Cool":
        print("Picture_ColorTemperature记忆Cool模式")
    else:
        print("Picture_ColorTemperature未记忆Cool模式，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature未记忆Cool模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature未记忆Cool模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "Warm":
        print("Picture_ColorTemperature切换Warm模式成功")
    else:
        print("Picture_ColorTemperature切换Warm模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature切换Warm模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature切换Warm模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    #Standard
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "Warm":
        print("Picture_ColorTemperature记忆Warm模式")
    else:
        print("Picture_ColorTemperature未记忆Warm模式，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature未记忆Warm模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature未记忆Warm模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "Standard":
        print("Picture_ColorTemperature切换Standard模式成功")
    else:
        print("Picture_ColorTemperature切换Standard模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature切换Standard模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','错误','Picture_ColorTemperature切换Standard模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','正确','Picture_ColorTemperature默认Standard模式且切换后可记忆'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature默认Standard模式且切换后可记忆','正确','Picture_ColorTemperature默认Standard模式且切换后可记忆'])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    for i in range(2):
        d.press("back")
        time.sleep(1)
    try:
        os.remove('D:\\auto_test_log\Picture_ColorTemperature.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
        
##Color Temperature参数修改
def Picture_ColorTemperature_modify():
    os.system("start d:\logcat.bat Picture_ColorTemperature_User")
    global sheet
    global workbook
    global file_path_Picture
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    for i in range(3):
        d.press("center")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "User":
        print("Picture_ColorTemperature切换User模式成功")
    else:
        print("Picture_ColorTemperature切换User模式失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature切换User模式失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature切换User模式失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    RedGain = 1
    GreenGain = 0
    BlueGain = 4
    RedOffset = 0
    GreenOffset = 0
    BlueOffset = 0
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedGain and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenGain and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueGain and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedOffset and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenOffset and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueOffset:
        print("Picture_ColorTemperature-User模式初始参数正确")
    else:
        print("Picture_ColorTemperature-User模式初始参数错误，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User模式初始参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User模式初始参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #RedGain
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedGain - 5:
        print("Picture_ColorTemperature-User-RedGain修改正确")
        RedGain = RedGain - 5
    else:
        print("Picture_ColorTemperature-User-RedGain修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedGain修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedGain修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #GreenGain
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenGain - 5:
        print("Picture_ColorTemperature-User-GreenGain修改正确")
        GreenGain = GreenGain - 5
    else:
        print("Picture_ColorTemperature-User-GreenGain修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenGain修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenGain修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #BlueGain
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueGain - 5:
        print("Picture_ColorTemperature-User-BlueGain修改正确")
        BlueGain = BlueGain - 5
    else:
        print("Picture_ColorTemperature-User-BlueGain修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueGain修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueGain修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #RedOffset
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedOffset - 5:
        print("Picture_ColorTemperature-User-RedOffset修改正确")
        RedOffset = RedOffset - 5
    else:
        print("Picture_ColorTemperature-User-RedOffset修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedOffset修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedOffset修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #GreenOffset
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenOffset - 5:
        print("Picture_ColorTemperature-User-GreenOffset修改正确")
        GreenOffset = GreenOffset - 5
    else:
        print("Picture_ColorTemperature-User-GreenOffset修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenOffset修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenOffset修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #BlueOffset
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueOffset - 5:
        print("Picture_ColorTemperature-User-GreenOffset修改正确")
        BlueOffset = BlueOffset - 5
    else:
        print("Picture_ColorTemperature-User-BlueOffset修改错误")
        talbe.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueOffset修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueOffset修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    ##------------------------------------------------------------------##
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedGain and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenGain and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueGain and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedOffset and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenOffset and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueOffset:
        print("Picture_ColorTemperature-User模式记忆修改参数")
    else:
        print("Picture_ColorTemperature-User模式未记忆修改参数，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User模式未记忆修改参数'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User模式未记忆修改参数'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #RedGain
    d.press("down")
    time.sleep(1)
    for i in range(10):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedGain + 10:
        print("Picture_ColorTemperature-User-RedGain修改正确")
        RedGain = RedGain + 10
    else:
        print("Picture_ColorTemperature-User-RedGain修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedGain修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedGain修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #GreenGain
    d.press("down")
    time.sleep(1)
    for i in range(10):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenGain + 10:
        print("Picture_ColorTemperature-User-GreenGain修改正确")
        GreenGain = GreenGain + 10
    else:
        print("Picture_ColorTemperature-User-GreenGain修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenGain修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenGain修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #BlueGain
    d.press("down")
    time.sleep(1)
    for i in range(10):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueGain + 10:
        print("Picture_ColorTemperature-User-BlueGain修改正确")
        BlueGain = BlueGain + 10
    else:
        print("Picture_ColorTemperature-User-BlueGain修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueGain修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueGain修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #RedOffset
    d.press("down")
    time.sleep(1)
    for i in range(10):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedOffset + 10:
        print("Picture_ColorTemperature-User-RedOffset修改正确")
        RedOffset = RedOffset + 10
    else:
        print("Picture_ColorTemperature-User-RedOffset修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedOffset修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedOffset修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #GreenOffset
    d.press("down")
    time.sleep(1)
    for i in range(10):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenOffset + 10:
        print("Picture_ColorTemperature-User-GreenOffset修改正确")
        GreenOffset = GreenOffset + 10
    else:
        print("Picture_ColorTemperature-User-GreenOffset修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenOffset修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenOffset修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #BlueOffset
    d.press("down")
    time.sleep(1)
    for i in range(10):
        d.press("right")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueOffset + 10:
        print("Picture_ColorTemperature-User-GreenOffset修改正确")
        BlueOffset = BlueOffset + 10
    else:
        print("Picture_ColorTemperature-User-BlueOffset修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueOffset修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueOffset修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    ##-------------------------------------------------##
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedGain and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenGain and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueGain and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedOffset and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenOffset and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueOffset:
        print("Picture_ColorTemperature-User模式记忆修改参数正确")
    else:
        print("Picture_ColorTemperature-User模式记忆修改参数错误，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User模式记忆修改参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User模式记忆修改参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #RedGain
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedGain - 5:
        print("Picture_ColorTemperature-User-RedGain修改正确")
        RedGain = RedGain - 5
    else:
        print("Picture_ColorTemperature-User-RedGain修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedGain修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedGain修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #GreenGain
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenGain - 5:
        print("Picture_ColorTemperature-User-GreenGain修改正确")
        GreenGain = GreenGain - 5
    else:
        print("Picture_ColorTemperature-User-GreenGain修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenGain修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenGain修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #BlueGain
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueGain - 5:
        print("Picture_ColorTemperature-User-BlueGain修改正确")
        BlueGain = BlueGain - 5
    else:
        print("Picture_ColorTemperature-User-BlueGain修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueGain修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueGain修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #RedOffset
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == RedOffset - 5:
        print("Picture_ColorTemperature-User-RedOffset修改正确")
        RedOffset = RedOffset - 5
    else:
        print("Picture_ColorTemperature-User-RedOffset修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedOffset修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-RedOffset修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #GreenOffset
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == GreenOffset - 5:
        print("Picture_ColorTemperature-User-GreenOffset修改正确")
        GreenOffset = GreenOffset - 5
    else:
        print("Picture_ColorTemperature-User-GreenOffset修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenOffset修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-GreenOffset修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    #BlueOffset
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == BlueOffset - 5:
        print("Picture_ColorTemperature-User-GreenOffset修改正确")
        BlueOffset = BlueOffset - 5
    else:
        print("Picture_ColorTemperature-User-BlueOffset修改错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueOffset修改错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-User-BlueOffset修改错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','正确','Picture_ColorTemperature的User模式下初始参数及修改参数正确'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的User模式下初始参数及修改参数正确','正确','Picture_ColorTemperature的User模式下初始参数及修改参数正确'])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Picture_ColorTemperature_User.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##Color Temperature任意模式下修改参数切换到User模式
def Picture_ColorTemperature_modify_to_user():
    os.system("start d:\logcat.bat Picture_ColorTemperature_")
    global sheet
    global workbook
    global file_path_Picture
    ##cool
    into_Picture()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == -1 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 32 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0:
        print("Picture_ColorTemperature-Cool模式初始参数正确")
    else:
        print("Picture_ColorTemperature-Cool模式初始参数错误，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-Cool模式初始参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-Cool模式初始参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("down")
    time.sleep(1)
    d.press("left")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "User":
        print("Picutre_ColorTemperature非User模式下修改参数自动切换到User模式")
    else:
        print("Picutre_ColorTemperature非User模式下修改参数未自动切换到User模式")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picutre_ColorTemperature非User模式下修改参数未自动切换到User模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picutre_ColorTemperature非User模式下修改参数未自动切换到User模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == -2 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 32 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0:
        print("Picture_ColorTemperature非User模式下修改参数自动切换到User模式，并保留参数")
    else:
        print("Picture_ColorTemperature非User模式下修改参数自动切换到User模式，未保留参数，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature非User模式下修改参数自动切换到User模式，未保留参数'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature非User模式下修改参数自动切换到User模式，未保留参数'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    ##---------------------------------------------------------##Standard
    d.press("up")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 1 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 4 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0:
        print("Picture_ColorTemperature-Standard模式初始参数正确")
    else:
        print("Picture_ColorTemperature-Standard模式初始参数错误，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-Standard模式初始参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-Standard模式初始参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("down")
    time.sleep(1)
    d.press("left")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "User":
        print("Picutre_ColorTemperature非User模式下修改参数自动切换到User模式")
    else:
        print("Picutre_ColorTemperature非User模式下修改参数未自动切换到User模式")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picutre_ColorTemperature非User模式下修改参数未自动切换到User模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picutre_ColorTemperature非User模式下修改参数未自动切换到User模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 4 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0:
        print("Picture_ColorTemperature非User模式下修改参数自动切换到User模式，并保留参数")
    else:
        print("Picture_ColorTemperature非User模式下修改参数自动切换到User模式，未保留参数，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature非User模式下修改参数自动切换到User模式，未保留参数'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature非User模式下修改参数自动切换到User模式，未保留参数'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    ##-------------------------------------------------------------##Warm
    d.press("up")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0:
        print("Picture_ColorTemperature-Warm模式初始参数正确")
    else:
        print("Picture_ColorTemperature-Warm模式初始参数错误，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-Warm模式初始参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature-Warm模式初始参数错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("down")
    time.sleep(1)
    d.press("left")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "User":
        print("Picutre_ColorTemperature非User模式下修改参数自动切换到User模式")
    else:
        print("Picutre_ColorTemperature非User模式下修改参数未自动切换到User模式")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picutre_ColorTemperature非User模式下修改参数未自动切换到User模式'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picutre_ColorTemperature非User模式下修改参数未自动切换到User模式'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == -1 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView').get_text()) == 0:
        print("Picture_ColorTemperature非User模式下修改参数自动切换到User模式，并保留参数")
    else:
        print("Picture_ColorTemperature非User模式下修改参数自动切换到User模式，未保留参数，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature非User模式下修改参数自动切换到User模式，未保留参数'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','错误','Picture_ColorTemperature非User模式下修改参数自动切换到User模式，未保留参数'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','正确','Picture_ColorTemperature的切换模式下初始参数及修改参数正确'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_ColorTemperature的切换模式下初始参数及修改参数正确','正确','Picture_ColorTemperature的切换模式下初始参数及修改参数正确'])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    d.press("back")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    try:
        os.remove('D:\\auto_test_log\Picture_ColorTemperature_.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
## 20 Point White Balance Correction参数修改
def Picture_PWBC():
    os.system("start d:\logcat.bat Picture_PWBC")
    global sheet
    global workbook
    global file_path_Picture
    into_Picture_Mode_User()
    into_Picture()
    for i in range(10):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == False:
        print("Picture_PWBC_Enable默认关闭")
    else:
        print("Picture_PWBC_Enable默认开启，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Enable默认开启'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Enable默认开启'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    
    for i in range(5):
        d.press("down")
        time.sleep(1)
    isFocused = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]').info["focused"]
    print(isFocused)
    if isFocused == "true":
        print("Picture_PWBC_Enable关闭状态不可修改参数")
    else:
        print("Picture_PWBC_Enable关闭状态可修改参数，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Enable关闭状态可修改参数'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Enable关闭状态可修改参数'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
        
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == True:
        print("Picture_PWBC_Enable开启成功")
    else:
        print("Picture_PWBC_Enable开启失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Enable开启失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Enable开启失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    Gain = d(resourceId="android:id/summary").get_text()
    Red = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text())
    Green = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text())
    Blue = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text())
    if Gain == "5％" and Red == 50 and Green == 50 and Blue == 50:
        print("Picture_PWBC参数初始值正确")
    else:
        print("Picture_PWBC参数初始值错误，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC参数初始值错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC参数初始值错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    print(d(resourceId="android:id/summary").get_text())
    if d(resourceId="android:id/summary").get_text() == "10％":
        print("Picture_PWBC_Gain修改成功")
    else:
        print("Picture_PWBC_Gain修改失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Gain修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Gain修改失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        for j in range(5):
            d.press("left")
            time.sleep(1)
        d.press("back")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Red - 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Green - 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Blue - 5:
        print("Picture_PWBC_Red、Green、Blue修改成功")
    else:
        print("Picture_PWBC_Red、Green、Blue修改失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Red、Green、Blue修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Red、Green、Blue修改失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    
    for i in range(3):
        d.press("up")
        time.sleep(1)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Red and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Green and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Blue:
        print("Picture_PWBC_Gain不同数值修改参数之间互不影响")
    else:
        print("Picture_PWBC_Gain不同数值修改参数之间有影响，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Gain不同数值修改参数之间有影响'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Gain不同数值修改参数之间有影响'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    
    for i in range(2):
        d.press("back")
        time.sleep(1)
    into_Picture()
    for i in range(10):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == True:
        print("Picture_PWBC_Enable记忆修改")
    else:
        print("Picture_PWBC_Enable未记忆修改，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Enable未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Enable未记忆修改'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    if d(resourceId="android:id/summary").get_text() == "10％":
        print("Picture_PWBC_Gain记忆修改")
    else:
        print("Picture_PWBC_Gain未记忆修改，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Gain未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Gain未记忆修改'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Red - 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Green - 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Blue - 5:
        print("Picture_PWBC_Red、Green、Blue记忆修改")
    else:
        print("Picture_PWBC_Red、Green、Blue未记忆修改，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Red、Green、Blue未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Red、Green、Blue未记忆修改'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("down")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        for j in range(10):
            d.press("right")
            time.sleep(1)
        d.press("back")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Red + 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Green + 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Blue + 5:
        print("Picture_PWBC_Red、Green、Blue修改成功")
    else:
        print("Picture_PWBC_Red、Green、Blue修改失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Red、Green、Blue修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Red、Green、Blue修改失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        for j in range(5):
            d.press("left")
            time.sleep(1)
        d.press("back")
        time.sleep(1)
        d.press("up")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Red and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Green and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Blue:
        print("Picture_PWBC_Red、Green、Blue修改成功")
    else:
        print("Picture_PWBC_Red、Green、Blue修改失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Red、Green、Blue修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Red、Green、Blue修改失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
        
    for i in range(2):
        d.press("center")
        time.sleep(1)
    if d(resourceId="android:id/summary").get_text() == "5％":
        print("Picture_PWBC_Gain修改成功")
    else:
        print("Picture_PWBC_Gain修改成功，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Gain修改成功'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Gain修改成功'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("up")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == False:
        print("Picture_PWBC_Enable关闭成功")
    else:
        print("Picture_PWBC_Enable关闭失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Enable默认开启'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_PWBC_Enable默认开启'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','正确','Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','正确','Picture_PWBC_Enable默认关闭，关闭不可修改参数，开启可修改可记忆'])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Picture_PWBC.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##Advanced Settings参数修改
def Picture_AdvancedSettings():
    os.system("start d:\logcat.bat Picture_AdvancedSettings")
    global sheet
    global workbook
    global file_path_Picture
    DNR = "Medium"
    Adaptiv_Luma_Control = "Medium"
    Local_Contrast_Control = "Medium"
    MPEG_NR = "Medium"
    DI_Film_Mode = "Auto"
    Color_Space = "Off"
    arr1 = ['Off','Low','Medium','High']
    arr2 = ['Off','Auto']
    arr3 = ['Auto','DCI-P3','BT.2020','sRGB','BT.709','Adobe RGB','Off']
    arr4 = ['Off','Low','Medium','Strong']
    arr1_len = 4
    arr2_len = 2
    arr3_len = 7
    arr4_len = 4
    
    into_Picture_Mode_User()
    into_Picture()
    for i in range(11):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == DNR and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == Adaptiv_Luma_Control and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == Local_Contrast_Control and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == MPEG_NR and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == DI_Film_Mode and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == Color_Space:
        print("Advanced Settings各参数初始状态正确")
    else:
        print("Advanced Settings各参数初始状态错误，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Advanced Settings参数默认正确，修改后可记忆','错误','Advanced Settings各参数初始状态错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Advanced Settings参数默认正确，修改后可记忆','错误','Advanced Settings各参数初始状态错误'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(8):
        d.press("center")
        time.sleep(1)
        for j in range(i % arr1_len):
            d.press("down")
            time.sleep(1)
        d.press("center")
        time.sleep(1)
        
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        for j in range(i % arr4_len):
            d.press("down")
            time.sleep(1)
        d.press("center")
        time.sleep(1)
        
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        for j in range(i % arr1_len):
            d.press("down")
            time.sleep(1)
        d.press("center")
        time.sleep(1)
        
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        for j in range(i % arr1_len):
            d.press("down")
            time.sleep(1)
        d.press("center")
        time.sleep(1)
        
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        for j in range(i % arr2_len):
            d.press("down")
            time.sleep(1)
        d.press("center")
        time.sleep(1)
            
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        for j in range(i % arr3_len):
            d.press("down")
            time.sleep(1)
        d.press("center")
        time.sleep(1)
        
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text())
        print(arr1[i % arr1_len])
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text())
        print(arr4[i % arr4_len])
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text())
        print(arr1[i % arr1_len])
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text())
        print(arr1[i % arr1_len])
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text())
        print(arr2[i % arr2_len])
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text())
        print(arr3[i % arr3_len])
        if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr1[i % arr1_len] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr4[i % arr4_len] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr1[i % arr1_len] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr1[i % arr1_len] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr2[i % arr2_len] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr3[i % arr3_len]:
            print("Advanced Settings参数修改成功")
        else:
            print("Advanced Settings参数修改失败，错误")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Advanced Settings参数默认正确，修改后可记忆','错误','Advanced Settings参数修改失败'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Advanced Settings参数默认正确，修改后可记忆','错误','Advanced Settings参数修改失败'])
            workbook.save(file_path_Picture)
            os.system("adb shell kill `pgrep logcat`")
            return
        if i == 0:
            d.press("back")
            time.sleep(1)
            d.press("back")
            time.sleep(1)
            into_Picture()
            for j in range(11):
                d.press("down")
                time.sleep(1)
            d.press("center")
            time.sleep(1)
            if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr1[i % arr1_len] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr4[i % arr4_len] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr1[i % arr1_len] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr1[i % arr1_len] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr2[i % arr2_len] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == arr3[i % arr3_len]:
                print("Advanced Settings参数记忆修改")
            else:
                print("Advanced Settings参数未记忆修改，错误")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Advanced Settings参数默认正确，修改后可记忆','错误','Advanced Settings参数未记忆修改'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Advanced Settings参数默认正确，修改后可记忆','错误','Advanced Settings参数未记忆修改'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
        while(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]').info["focused"] != "true"):
            d.press("up")
            time.sleep(1)
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Advanced Settings参数默认正确，修改后可记忆','正确','Advanced Settings参数默认正确，修改后可记忆'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Advanced Settings参数默认正确，修改后可记忆','正确','Advanced Settings参数默认正确，修改后可记忆'])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    d.press("center")
    time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        for j in range(2):
            d.press("down")
            time.sleep(1)
        d.press("center")
        time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(6):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(2):
        d.press("back")
        time.sleep(1)
    try:
        os.remove('D:\\auto_test_log\Picture_AdvancedSettings.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)

##Color Tuner参数修改
def Picture_ColorTuner():
    os.system("start d:\logcat.bat Picture_ColorTuner")
    global sheet
    global workbook
    global file_path_Picture
    into_Picture_Mode_User()
    into_Picture()
    for i in range(12):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == False:
        print("Picture_Color Tuner_Enable默认关闭")
    else:
        print("Picture_Color Tuner_Enable默认开启，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner默认开启'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner默认开启'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("down")
        time.sleep(1)
    isFocused = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]').info["focused"]
    print(isFocused)
    if isFocused == "true":
        print("Picture_Color Tuner_Enable关闭状态不可修改参数")
    else:
        print("Picture_Color Tuner_Enable关闭状态可修改参数，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Enable关闭状态可修改参数'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Enable关闭状态可修改参数'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == True:
        print("Picture_Color Tuner_Enable开启成功")
    else:
        print("Picture_Color Tuner_Enable开启失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Enable开启失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Enable开启失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return

    for i in range(3):
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        for j in range(7):
            for k in range(5):
                d.press("left")
                time.sleep(1)
            d.press("down")
            time.sleep(1)
        d.press("back")
        time.sleep(1)
        for j in range(6):
            d.press("up")
            time.sleep(1)
        if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45":
            if i == 0:
                print("Picture_Color Tuner_HUE参数修改成功")
            elif i == 1:
                print("Picture_Color Tuner_Saturation参数修改成功")
            else:
                print("Picture_Color Tuner_Brightness参数修改成功")
        else:
            if i == 0:
                print("Picture_Color Tuner_HUE参数修改失败")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_HUE参数修改失败'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_HUE参数修改失败'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
            elif i == 1:
                print("Picture_Color Tuner_Saturation参数修改失败")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Saturation参数修改失败'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Saturation参数修改失败'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
            else:
                print("Picture_Color Tuner_Brightness参数修改失败")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Brightness参数修改失败'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Brightness参数修改失败'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
        d.press("back")
        time.sleep(1)
        
    for i in range(2):
        d.press("back")
        time.sleep(1)  
    
    ##--------------45-55---------------##
    into_Picture()
    for i in range(3):
        d.press("up")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == True:
        print("Picture_Color Tuner_Enable记忆开启")
    else:
        print("Picture_Color Tuner_Enable未记忆开启，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Enable未记忆开启'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Enable未记忆开启'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
        
    for i in range(3):
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "45":
            if i == 0:
                print("Picture_Color Tuner_HUE参数记忆修改")
            elif i == 1:
                print("Picture_Color Tuner_Saturation参数记忆修改")
            else:
                print("Picture_Color Tuner_Brightness参数记忆修改")
        else:
            if i == 0:
                print("Picture_Color Tuner_HUE参数未记忆修改")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_HUE参数未记忆修改'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_HUE参数未记忆修改'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
            elif i == 1:
                print("Picture_Color Tuner_Saturation参数未记忆修改")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Saturation参数未记忆修改'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Saturation参数未记忆修改'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
            else:
                print("Picture_Color Tuner_Brightness参数未记忆修改")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Brightness参数未记忆修改'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Brightness参数未记忆修改'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
        
        for j in range(7):
            for k in range(10):
                d.press("right")
                time.sleep(1)
            d.press("down")
            time.sleep(1)
        d.press("back")
        time.sleep(1)
        for j in range(6):
            d.press("up")
            time.sleep(1)
        if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55":
            if i == 0:
                print("Picture_Color Tuner_HUE参数修改成功")
            elif i == 1:
                print("Picture_Color Tuner_Saturation参数修改成功")
            else:
                print("Picture_Color Tuner_Brightness参数修改成功")
        else:
            if i == 0:
                print("Picture_Color Tuner_HUE参数修改失败")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_HUE参数修改失败'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_HUE参数修改失败'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
            elif i == 1:
                print("Picture_Color Tuner_Saturation参数修改失败")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Saturation参数修改失败'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Saturation参数修改失败'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
            else:
                print("Picture_Color Tuner_Brightness参数修改失败")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Brightness参数修改失败'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Brightness参数修改失败'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
        d.press("back")
        time.sleep(1)

    for i in range(2):
        d.press("back")
        time.sleep(1) 
    
    ##-----------------55-50-------------------##
    into_Picture()
    for i in range(3):
        d.press("up")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == True:
        print("Picture_Color Tuner_Enable记忆开启")
    else:
        print("Picture_Color Tuner_Enable未记忆开启，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Enable未记忆开启'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Enable未记忆开启'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
        
    for i in range(3):
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        
        if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "55":
            if i == 0:
                print("Picture_Color Tuner_HUE参数记忆修改")
            elif i == 1:
                print("Picture_Color Tuner_Saturation参数记忆修改")
            else:
                print("Picture_Color Tuner_Brightness参数记忆修改")
        else:
            if i == 0:
                print("Picture_Color Tuner_HUE参数未记忆修改")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_HUE参数未记忆修改'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_HUE参数未记忆修改'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
            elif i == 1:
                print("Picture_Color Tuner_Saturation参数未记忆修改")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Saturation参数未记忆修改'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Saturation参数未记忆修改'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
            else:
                print("Picture_Color Tuner_Brightness参数未记忆修改")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Brightness参数未记忆修改'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Brightness参数未记忆修改'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
        
        for j in range(7):
            for k in range(5):
                d.press("left")
                time.sleep(1)
            d.press("down")
            time.sleep(1)
        d.press("back")
        time.sleep(1)
        for j in range(6):
            d.press("up")
            time.sleep(1)
        if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50":
            if i == 0:
                print("Picture_Color Tuner_HUE参数修改成功")
            elif i == 1:
                print("Picture_Color Tuner_Saturation参数修改成功")
            else:
                print("Picture_Color Tuner_Brightness参数修改成功")
        else:
            if i == 0:
                print("Picture_Color Tuner_HUE参数修改失败")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_HUE参数修改失败'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_HUE参数修改失败'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
            elif i == 1:
                print("Picture_Color Tuner_Saturation参数修改失败")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Saturation参数修改失败'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Saturation参数修改失败'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
            else:
                print("Picture_Color Tuner_Brightness参数修改失败")
                table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Brightness参数修改失败'))
                sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color Tuner_Enable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Brightness参数修改失败'])
                workbook.save(file_path_Picture)
                os.system("adb shell kill `pgrep logcat`")
                return
        d.press("back")
        time.sleep(1) 
    for i in range(3):
        d.press("up")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == False:
        print("Picture_Color Tuner_Enable关闭成功")
    else:
        print("Picture_Color Tuner_Enable关闭失败，错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Enable关闭失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆','错误','Picture_Color Tuner_Enable关闭失败'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆','正确','Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆','正确','Picture_Color TunerEnable默认关闭，关闭不可修改参数，开启可修改可记忆'])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    for i in range(2):
        d.press("back")
        time.sleep(1)
    try:
        os.remove('D:\\auto_test_log\Picture_ColorTuner.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
        
##Picture Reset
def Picture_Reset():
    os.system("start d:\logcat.bat Picture_Reset")
    global sheet
    global workbook
    global file_path_Picture
    into_Picture_Mode_User()
    into_Picture()
    
    d.press("down")
    time.sleep(1)
    for i in range(6):
        d.press("down")
        time.sleep(1)
        d.press("left")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    for i in range(2):
        d.press("center")
        time.sleep(1)
        
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(6):
        d.press("down")
        time.sleep(1)
        d.press("left")
        time.sleep(1)
    for i in range(2):
        d.press("back")
        time.sleep(1)
    
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if not isChecked:
        d.press("center")
        time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
        d.press("left")
        time.sleep(1)
    for i in range(2):
        d.press("back")
        time.sleep(1)
    
    d.press("down")
    time.sleep(1)
    for i in range(3):
        d.press("center")
        time.sleep(1)
    d.press("back")
    time.sleep(1)
    
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if not isChecked:
        d.press("center")
        time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(7):
        d.press("left")
        time.sleep(1)
        d.press("down")
        time.sleep(1)
    for i in range(3):
        d.press("back")
        time.sleep(1)
    
    for i in range(2):
        d.press("down")
        time.sleep(1)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    time.sleep(15)
    
    d.press("down")
    time.sleep(1)
    
    for i in range(3):
        d.press("down")
        time.sleep(1)
    
    flag = 1
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "Energy Saving" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "100" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "0" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "10":
        flag = 1
        print("Mode")
    else:
        flag = 0
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture模式或模式参数'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture模式或模式参数'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return 
        
    for i in range(4):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "Middle":
        flag = 1
        print("Gamma")
    else:
        flag = 0
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture_Gamma'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture_Gamma'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return 
    
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "Standard" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "1" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "0" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "4" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "0" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "0" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "0":
        flag = 1
        print("Color Temperature")
    else:
        flag = 0
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture_Color Temperature'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture_Color Temperature'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return 
    d.press("back")
    time.sleep(1)
    
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked == False and d.xpath('//*[@resource-id="android:id/summary"]').get_text() == "5％" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50":
        flag = 1
        print("PWBC")
    else:
        print(isChecked)
        print(d.xpath('//*[@resource-id="android:id/summary"]').get_text())
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text())
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text())
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text())
        flag = 0
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture_PWBC'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture_PWBC'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return 
    d.press("back")
    time.sleep(1)
    
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Medium" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Medium" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Medium" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Medium" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Auto" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Off":
        flag = 1
        print("Advanced Settings")
    else:
        flag = 0
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture_Advanced Settings'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture_Advanced Settings'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return 
    d.press("back")
    time.sleep(1)
    
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if not isChecked == True:
        d.press("center")
        time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50":
        flag = 1
        print("Color Tuner")
    else:
        flag = 0
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture_ColorTuner'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','错误','未重置Picture_ColorTuner'])
        workbook.save(file_path_Picture)
        os.system("adb shell kill `pgrep logcat`")
        return 
    
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','正确','Picture Reset重置所有数据'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),'Picture Reset重置所有数据','正确','Picture Reset重置所有数据'])
    workbook.save(file_path_Picture)
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Picture_Reset.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
        
##Picture模式图像对比
def Picture_Mode_compare_screen():
    d.press("home")
    time.sleep(1)
    d.shell("am start -n com.smartdevice.livetv/skyworth.skyworthlivetv.osd.ui.mainActivity.LiveTvScreenActivity")
    time.sleep(5)
    d.shell("input keyevent 13")
    time.sleep(5)
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(2)
    for i in range(3):
        d.press("center")
        time.sleep(1)
    for i in range(4):
        d.press("back")
        time.sleep(1)
    
    cap = cv2.VideoCapture(1)
    if not cap.isOpened():
        print("开启摄像头失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture切换模式效果明显",'错误',"开启摄像头失败"))
        return
    time.sleep(3)
    ret,frame = cap.read()
    cv2.imwrite('D:\\test\Picture_Vivid.jpg',frame)
    
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(2)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(4):
        d.press("back")
        time.sleep(1)

    time.sleep(3)
    ret,frame = cap.read()
    cv2.imwrite('D:\\test\Picture_Standard.jpg',frame)
    
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(2)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(4):
        d.press("back")
        time.sleep(1)
    
    time.sleep(3)
    ret,frame = cap.read()
    cv2.imwrite('D:\\test\Picture_Sport.jpg',frame)
    
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(2)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(4):
        d.press("back")
        time.sleep(1)
    
    time.sleep(3)
    ret,frame = cap.read()
    cv2.imwrite('D:\\test\Picture_Movie.jpg',frame)
    
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(2)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    for i in range(4):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(4):
        d.press("back")
        time.sleep(1)

    time.sleep(3)
    ret,frame = cap.read()
    cv2.imwrite('D:\\test\Picture_User.jpg',frame)
    
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(2)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    for i in range(5):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(4):
        d.press("back")
        time.sleep(1)
    
    time.sleep(3)
    ret,frame = cap.read()
    cv2.imwrite('D:\\test\Picture_Energy_Saving.jpg',frame)
    
    cap.release()
    cv2.destroyAllWindows()
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Vivid.jpg","D:\\test\Picture_Standard.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Vivid.jpg","D:\\test\Picture_Sport.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Vivid.jpg","D:\\test\Picture_Movie.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Vivid.jpg","D:\\test\Picture_User.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Vivid.jpg","D:\\test\Picture_Energy_Saving.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Standard.jpg","D:\\test\Picture_Sport.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Standard.jpg","D:\\test\Picture_Movie.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Standard.jpg","D:\\test\Picture_User.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Standard.jpg","D:\\test\Picture_Energy_Saving.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Sport.jpg","D:\\test\Picture_Movie.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Sport.jpg","D:\\test\Picture_User.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Sport.jpg","D:\\test\Picture_Energy_Saving.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Movie.jpg","D:\\test\Picture_User.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_Movie.jpg","D:\\test\Picture_Energy_Saving.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_User.jpg","D:\\test\Picture_Energy_Saving.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'正确',"Picture模式切换效果明显"))
    
##Picture cool temperature 模式图像对比
def Picture_Mode_cool_Temperature_compare_screen():
    d.press("home")
    time.sleep(1)
    d.shell("am start -n com.smartdevice.livetv/skyworth.skyworthlivetv.osd.ui.mainActivity.LiveTvScreenActivity")
    time.sleep(5)
    d.shell("input keyevent 13")
    time.sleep(5)
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(2)
    d.press("center")
    time.sleep(1)
    for i in range(8):
        d.press("down")
        time.sleep(1)
    for i in range(3):
        d.press("center")
        time.sleep(1)
    for i in range(5):
        d.press("back")
        time.sleep(1)
        
    cap = cv2.VideoCapture(1)
    if not cap.isOpened():
        print("开启摄像头失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture切换模式效果明显",'错误',"开启摄像头失败"))
        return
        
    time.sleep(3)
    ret,frame = cap.read()
    cv2.imwrite('D:\\test\Picture_CoolT_User.jpg',frame)
    
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(2)
    d.press("center")
    time.sleep(1)
    for i in range(8):
        d.press("down")
        time.sleep(1)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(5):
        d.press("back")
        time.sleep(1)
    time.sleep(3)
    ret,frame = cap.read()
    cv2.imwrite('D:\\test\Picture_CoolT_Cool.jpg',frame)

    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(2)
    d.press("center")
    time.sleep(1)
    for i in range(8):
        d.press("down")
        time.sleep(1)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(5):
        d.press("back")
        time.sleep(1)
    time.sleep(3)
    ret,frame = cap.read()
    cv2.imwrite('D:\\test\Picture_CoolT_Standard.jpg',frame)
    
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(2)
    d.press("center")
    time.sleep(1)
    for i in range(8):
        d.press("down")
        time.sleep(1)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(5):
        d.press("back")
        time.sleep(1)
    time.sleep(3)
    ret,frame = cap.read()
    cv2.imwrite('D:\\test\Picture_CoolT_Warm.jpg',frame)
    
    cap.release()
    cv2.destroyAllWindows()
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_CoolT_User.jpg","D:\\test\Picture_CoolT_Cool.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_CoolT_User.jpg","D:\\test\Picture_CoolT_Standard.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_CoolT_User.jpg","D:\\test\Picture_CoolT_Warm.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_CoolT_Cool.jpg","D:\\test\Picture_CoolT_Standard.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_CoolT_Cool.jpg","D:\\test\Picture_CoolT_Warm.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    
    m1,m2,m3 = runAllImageSimilaryFun("D:\\test\Picture_CoolT_Standard.jpg","D:\\test\Picture_CoolT_Warm.jpg")
    if int(m1) + int(m2) + int(m3) <= 10:
        print("Picture模式切换效果不明显")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),"Picture模式切换效果明显",'错误',"Picture模式切换效果不明显"))
        return
    

    
##进入Sound
def into_Sound():
    print("into_Sound")
    d.press("home")
    time.sleep(1)
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.SoundActivity")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    
##默认状态
def Sound_Default():
    os.system("start d:\logcat.bat Sound_Default")
    global sheet
    global workbook
    global file_path_Sound
    print("Sound_Default")
    into_Sound()
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked:
        print("Sound_DolbyAtmos默认打开")
    else:
        print("Sound_DolbyAtmos默认未打开")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','Sound_DolbyAtmos默认未打开'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','Sound_DolbyAtmos默认未打开'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    isEnabled_SoundMode = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]').info["enabled"]
    print(isEnabled_SoundMode)
    isEnabled_DolbySoundMode = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]').info["enabled"]
    isEnabled_Balance = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]').info["enabled"]
    isEnabled_Bass = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]').info["enabled"]
    isEnabled_Treble = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]').info["enabled"]
    isEnabled_DTSStudioSound = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]').info["enabled"]
    if isEnabled_SoundMode == "true" or isEnabled_DolbySoundMode == "false" or isEnabled_Balance == "true" or isEnabled_Bass == "true" or isEnabled_Treble == "true" or isEnabled_DTSStudioSound == "true":
        print("Sound_DolbyAtmos开启状态下出现不可选的参数可选")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','Sound_DolbyAtmos开启状态下出现不可选的参数可选'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','Sound_DolbyAtmos开启状态下出现不可选的参数可选'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Standard" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Standard" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "0" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50":
        print("部分参数初始化状态正确")
    else:
        print("部分参数初始化状态错误")
        table.insert('','end',value = (test_num, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','部分参数初始化状态错误'))
        sheet.append([test_num, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','部分参数初始化状态错误'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    
    for i in range(4):
        d.press("down")
        time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if isChecked:
        print("Sound_AutoVolume默认打开")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','Sound_AutoVolume默认打开'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','Sound_AutoVolume默认打开'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    isEnabled_AutoVolume = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]').info["enabled"]
    isEnabled_Speaker = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]').info["enabled"]
    isEnabled_AudioType = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]').info["enabled"]
    isEnabled_VisuallyImpaired = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]').info["enabled"]
    isEnabled_DialogueEnhancer = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]').info["enabled"]
    isEnabled_SoundReset = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]').info["enabled"]
    if isEnabled_AutoVolume == "true" or isEnabled_Speaker == "false" or isEnabled_AudioType == "false" or isEnabled_VisuallyImpaired == "true" or isEnabled_DialogueEnhancer == "false" or isEnabled_SoundReset == "false":
        print("Sound_DolbyAtmos开启状态下出现不可选的参数可选")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','Sound_DolbyAtmos开启状态下出现不可选的参数可选'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','Sound_DolbyAtmos开启状态下出现不可选的参数可选'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Auto" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Normal" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Off":
        print("部分参数初始化状态正确")
    else:
        print("部分参数初始化状态错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','部分参数初始化状态错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','错误','部分参数初始化状态错误'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','正确','Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确','正确','Sound_DolbyAtmos默认打开，相应参数不可修改，所有参数初始值正确'])
    workbook.save(file_path_Sound)   
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_Default.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)

##Dolby Atmos
def Sound_DolbyAtmos():
    os.system("start d:\logcat.bat Sound_DolbyAtmos")
    global sheet
    global workbook
    global file_path_Sound
    
    into_Sound()
    
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if not isChecked:
        d.press("center")
        time.sleep(1)
        if d(resourceId="android:id/switch_widget").info["checked"]:
            print("Sound_DolbyAtmos修改成功")
        else:
            print("Sound_DolbyAtmos修改失败")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','错误','Picture_DolbyAtmos修改失败'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','错误','Sound_DolbyAtmos修改失败'])
            workbook.save(file_path_Sound)
            os.system("adb shell kill `pgrep logcat`")
            return
    
    d.press("center")
    time.sleep(1)
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if not isChecked:
        print("Sound_DolbyAtmos修改成功")
    else:
        print("Sound_DolbyAtmos修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','错误','Picture_DolbyAtmos修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','错误','Sound_DolbyAtmos修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    isEnabled_SoundMode = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]').info["enabled"]
    print(isEnabled_SoundMode)
    isEnabled_DolbySoundMode = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]').info["enabled"]
    isEnabled_Balance = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]').info["enabled"]
    isEnabled_Bass = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]').info["enabled"]
    isEnabled_Treble = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]').info["enabled"]
    isEnabled_DTSStudioSound = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]').info["enabled"]
    if isEnabled_SoundMode == "false" or isEnabled_DolbySoundMode == "true" or isEnabled_Balance == "false" or isEnabled_Bass == "false" or isEnabled_Treble == "false" or isEnabled_DTSStudioSound == "flase":
        print("Sound_DolbyAtmos开启状态下出现不可选的参数可选")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','错误','Sound_DolbyAtmos开启状态下出现不可选的参数可选'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','错误','Sound_DolbyAtmos开启状态下出现不可选的参数可选'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(9):
        d.press("down")
        time.sleep(1)
    isEnabled_AutoVolume = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]').info["enabled"]
    isEnabled_Speaker = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]').info["enabled"]
    isEnabled_AudioType = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]').info["enabled"]
    isEnabled_VisuallyImpaired = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]').info["enabled"]
    isEnabled_DialogueEnhancer = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]').info["enabled"]
    isEnabled_SoundReset = d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[7]').info["enabled"]
    if isEnabled_AutoVolume == "false" or isEnabled_Speaker == "false" or isEnabled_AudioType == "false" or isEnabled_VisuallyImpaired == "true" or isEnabled_DialogueEnhancer == "false" or isEnabled_SoundReset == "false":
        print("Sound_DolbyAtmos开启状态下出现不可选的参数可选")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','错误','Sound_DolbyAtmos开启状态下出现不可选的参数可选'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','错误','Sound_DolbyAtmos开启状态下出现不可选的参数可选'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("back")
        time.sleep(1)
    into_Sound()
    isChecked = d(resourceId="android:id/switch_widget").info["checked"]
    if not isChecked:
        print("Picture_DolbyAtmos记忆修改")
    else:
        print("Picture_DolbyAtmos未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','错误','Picture_DolbyAtmos未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','错误','Picture_DolbyAtmos未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','正确','Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选','正确','Sound_DolbyAtmos可修改，且记忆修改，关闭时部分参数可选'])
    workbook.save(file_path_Sound)    
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_DolbyAtmos.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)

##Sound Mode
def Sound_SoundMode():
    os.system("start d:\logcat.bat Sound_SoundMode")
    global sheet
    global workbook
    global file_path_Sound

    ##----------------Music---------------------##
    into_Sound()
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Music":
        print("Sound_SoundMode修改成功")
    else:
        print("Sound_SoundMode修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    into_Sound()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Music":
        print("Sound_SoundMode记忆修改")
    else:
        print("Sound_SoundMode未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    ##------------------Sport------------------##
    d.press("center")
    time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Sport":
        print("Sound_SoundMode修改成功")
    else:
        print("Sound_SoundMode修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    into_Sound()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Sport":
        print("Sound_SoundMode记忆修改")
    else:
        print("Sound_SoundMode未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    ##--------------------Movie---------------##
    d.press("center")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Movie":
        print("Sound_SoundMode修改成功")
    else:
        print("Sound_SoundMode修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    into_Sound()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Movie":
        print("Sound_SoundMode记忆修改")
    else:
        print("Sound_SoundMode未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    ##-----------------Standard---------------##
    for i in range(2):
        d.press("center")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Standard":
        print("Sound_SoundMode修改成功")
    else:
        print("Sound_SoundMode修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    into_Sound()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Standard":
        print("Sound_SoundMode记忆修改")
    else:
        print("Sound_SoundMode未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbyAtmos可修改，且记忆修改','错误','Sound_SoundMode未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Sound_DolbyAtmos可修改，且记忆修改','正确','Sound_DolbyAtmos可修改，且记忆修改'))
    sheet.append([test_num, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Sound_DolbyAtmos可修改，且记忆修改','正确','Sound_DolbyAtmos可修改，且记忆修改'])
    workbook.save(file_path_Sound)    
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_SoundMode.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    d.press("home")
    time.sleep(1)

##Dolby Sound Mode
def Sound_DolbySoundMode():
    os.system("start d:\logcat.bat Sound_DolbySoundMode")
    global sheet
    global workbook
    global file_path_Sound
    into_Sound()
    if not d(resourceId="android:id/switch_widget").info["checked"]:
        d.press("center")
        time.sleep(1)
        if d(resourceId="android:id/switch_widget").info["checked"]:
            print("Sound_DolbyAtmos打开成功")
        else:
            print("Sound_DolbyAtmos打开失败")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbyAtmos打开失败'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbyAtmos打开失败'])
            workbook.save(file_path_Sound)
            os.system("adb shell kill `pgrep logcat`")
            return
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Music":
        print("Sound_DolbySoundMode修改成功")
    else:
        print("Sound_DolbySoundMode修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("back")
        time.sleep(1)
    into_Sound()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Music":
        print("Sound_DolbySoundMode记忆修改")
    else:
        print("Sound_DolbySoundMode未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    ##------------------Sport------------------##
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Sport":
        print("Sound_DolbySoundMode修改成功")
    else:
        print("Sound_DolbySoundMode修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("back")
        time.sleep(1)
    into_Sound()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Sport":
        print("Sound_DolbySoundMode记忆修改")
    else:
        print("Sound_DolbySoundMode未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    ##--------------------Movie---------------##
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Movie":
        print("Sound_DolbySoundMode修改成功")
    else:
        print("Sound_DolbySoundMode修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("back")
        time.sleep(1)
    into_Sound()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Movie":
        print("Sound_DolbySoundMode记忆修改")
    else:
        print("Sound_DolbySoundMode未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    ##-----------------Standard---------------##
    d.press("down")
    time.sleep(1)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Standard":
        print("Sound_DolbySoundMode修改成功")
    else:
        print("Sound_DolbySoundMode修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("back")
        time.sleep(1)
    into_Sound()
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Standard":
        print("Sound_DolbySoundMode记忆修改")
    else:
        print("Sound_DolbySoundMode未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','错误','Sound_DolbySoundMode未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','正确','Sound_DolbySoundMode可修改，且记忆修改'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DolbySoundMode可修改，且记忆修改','正确','Sound_DolbySoundMode可修改，且记忆修改'])
    workbook.save(file_path_Sound)
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_DolbySoundMode.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    for i in range(3):
        d.press("back")
        time.sleep(1)
    
##Balance、Bass、Treble
def Sound_Balance_Bass_Treble():
    os.system("start d:\logcat.bat Sound_Balance_Bass_Treble")
    global sheet
    global workbook
    global file_path_Sound
    into_Sound()

    Balance = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text())
    Bass = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text())
    Treble = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text())
    print(Balance)
    for i in range(3):
        d.press("down")
        time.sleep(1)
        for j in range(5):
            d.press("left")
            time.sleep(1)
    print(int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()))
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Balance - 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Bass - 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Treble - 5:
        print("Balance、Bass、Treble修改成功")
    else:
        print("Balance、Bass、Treble修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','错误','Balance、Bass、Treble修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','错误','Balance、Bass、Treble修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    
    into_Sound()
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Balance - 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Bass - 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Treble - 5:
        print("Balance、Bass、Treble记忆修改")
    else:
        print("Balance、Bass、Treble未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','错误','Balance、Bass、Treble未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','错误','Balance、Bass、Treble未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("down")
        time.sleep(1)
        for j in range(10):
            d.press("right")
            time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Balance + 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Bass + 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Treble + 5:
        print("Balance、Bass、Treble修改成功")
    else:
        print("Balance、Bass、Treble修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','错误','Balance、Bass、Treble修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','错误','Balance、Bass、Treble修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
        
    into_Sound()
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Balance + 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Bass + 5 and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Treble + 5:
        print("Balance、Bass、Treble记忆修改")
    else:
        print("Balance、Bass、Treble未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','错误','Balance、Bass、Treble未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','错误','Balance、Bass、Treble未记忆修改'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("down")
        time.sleep(1)
        for j in range(5):
            d.press("left")
            time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Balance and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Bass and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Treble:
        print("Balance、Bass、Treble修改成功")
    else:
        print("Balance、Bass、Treble修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','错误','Balance、Bass、Treble修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','错误','Balance、Bass、Treble修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','正确','Sound_Balance、Bass、Treble可修改，且记忆修改'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Balance、Bass、Treble可修改，且记忆修改','正确','Sound_Balance、Bass、Treble可修改，且记忆修改'])
    workbook.save(file_path_Sound)
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_Balance_Bass_Treble.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    for i in range(2):
        d.press("back")
        time.sleep(1)
    
##DTS Studio Sound
def Sound_DTS():
    os.system("start d:\logcat.bat Sound_DTS")
    global sheet
    global workbook
    global file_path_Sound
    into_Sound()
    for i in range(4):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if (not d(resourceId="android:id/switch_widget")[0].info["checked"]) and (not d(resourceId="android:id/switch_widget")[1].info["checked"]) and (not d(resourceId="android:id/switch_widget")[2].info["checked"]):
        print("Sound_DTC_Enable默认关闭")
    else:
        print("Sound_DTC_Enable默认未关闭")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Sound_DTC_Enable默认未关闭'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Sound_DTC_Enable默认未关闭'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]').info["enabled"] == "false" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]').info["enabled"] == "false":
        print("Sound_DTS_Enable关闭状态Surround与TruVolume不可选中")
    else:
        print("Sound_DTS_Enable关闭状态Surround与TruVolume可选中")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Sound_DTS_Enable关闭状态Surround与TruVolume可选中'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Sound_DTS_Enable关闭状态Surround与TruVolume可选中'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("center")
        time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
    if d(resourceId="android:id/switch_widget")[0].info["checked"] and d(resourceId="android:id/switch_widget")[1].info["checked"] and d(resourceId="android:id/switch_widget")[2].info["checked"]:
        print("Enable、Surround、TruVolume修改成功")
    else:
        print("Enable、Surround、TruVolume修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume修改失败'])
        workbook.save(file_path_Sound)
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("back")
        time.sleep(1)
    
    into_Sound()
    d.press("center")
    time.sleep(1)
    if d(resourceId="android:id/switch_widget")[0].info["checked"] and d(resourceId="android:id/switch_widget")[1].info["checked"] and d(resourceId="android:id/switch_widget")[2].info["checked"]:
        print("Enable、Surround、TruVolume记忆修改")
    else:
        print("Enable、Surround、TruVolume未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume未记忆修改'])
        workbook.save(file_path_Sound)   
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    if (not d(resourceId="android:id/switch_widget")[0].info["checked"]) and (not d(resourceId="android:id/switch_widget")[1].info["checked"]) and (not d(resourceId="android:id/switch_widget")[2].info["checked"]):
        print("Enable、Surround、TruVolume修改成功")
    else:
        print("Enable、Surround、TruVolume修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("back")
        time.sleep(1)
    
    into_Sound()
    for i in range(4):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if (not d(resourceId="android:id/switch_widget")[0].info["checked"]) and (not d(resourceId="android:id/switch_widget")[1].info["checked"]) and (not d(resourceId="android:id/switch_widget")[2].info["checked"]):
        print("Enable、Surround、TruVolume记忆修改")
    else:
        print("Enable、Surround、TruVolume未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume未记忆修改'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("center")
        time.sleep(1)
    if d(resourceId="android:id/switch_widget")[0].info["checked"] and d(resourceId="android:id/switch_widget")[1].info["checked"] and d(resourceId="android:id/switch_widget")[2].info["checked"]:
        print("Enable、Surround、TruVolume修改成功")
    else:
        print("Enable、Surround、TruVolume修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
    for i in range(2):
        d.press("up")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if (not d(resourceId="android:id/switch_widget")[0].info["checked"]) and (not d(resourceId="android:id/switch_widget")[1].info["checked"]) and (not d(resourceId="android:id/switch_widget")[2].info["checked"]):
        print("Enable、Surround、TruVolume修改成功")
    else:
        print("Enable、Surround、TruVolume修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("back")
        time.sleep(1)
    
    into_Sound()
    for i in range(4):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if (not d(resourceId="android:id/switch_widget")[0].info["checked"]) and (not d(resourceId="android:id/switch_widget")[1].info["checked"]) and (not d(resourceId="android:id/switch_widget")[2].info["checked"]):
        print("Enable、Surround、TruVolume记忆修改")
    else:
        print("Enable、Surround、TruVolume未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume未记忆修改'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("center")
        time.sleep(1)
    if d(resourceId="android:id/switch_widget")[0].info["checked"] and (not d(resourceId="android:id/switch_widget")[1].info["checked"]) and (not d(resourceId="android:id/switch_widget")[2].info["checked"]):
        print("Enable、Surround、TruVolume记忆Enable开启时的设置的状态")
    else:
        print("Enable、Surround、TruVolume未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume记忆Enable开启时的设置的状态'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Enable、Surround、TruVolume记忆Enable开启时的设置的状态'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    if not d(resourceId="android:id/switch_widget")[0].info["checked"]:
        print("Sound_DTS_Enable关闭成功")
    else:
        print("Sound_DTS_Enable关闭失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Sound_DTS_Enable关闭失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DTSStudioSound可修改，且记忆修改','错误','Sound_DTS_Enable关闭失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Sound_DTSStudioSound可修改，且记忆修改','正确','Sound_DTSStudioSound可修改，且记忆修改'))
    sheet.append([test_num, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Sound_DTSStudioSound可修改，且记忆修改','正确','Sound_DTSStudioSound可修改，且记忆修改'])
    workbook.save(file_path_Sound) 
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_DTS.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    for i in range(3):
        d.press("back")
        time.sleep(1)

##Auto Voume
def Sound_AutoVolume():
    os.system("start d:\logcat.bat Sound_AutoVolume")
    global sheet
    global workbook
    global file_path_Sound
    into_Sound()
    for i in range(5):
        d.press("down")
        time.sleep(1)
    if not d(resourceId="android:id/switch_widget").info["checked"]:
        print("AutoVolume默认关闭")
    else:
        print("AutoVolume默认未关闭")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AutoVolume可修改，且记忆修改','错误','AutoVolume默认未关闭'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AutoVolume可修改，且记忆修改','错误','AutoVolume默认未关闭'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    if d(resourceId="android:id/switch_widget").info["checked"]:
        print("AutoVolume开启成功")
    else:
        print("AutoVolume开启失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AutoVolume可修改，且记忆修改','错误','AutoVolume开启失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AutoVolume可修改，且记忆修改','错误','AutoVolume开启失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    
    into_Sound()
    for i in range(5):
        d.press("down")
        time.sleep(1)
    if d(resourceId="android:id/switch_widget").info["checked"]:
        print("AutoVolume记忆修改")
    else:
        print("AutoVolume未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AutoVolume可修改，且记忆修改','错误','AutoVolume未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AutoVolume可修改，且记忆修改','错误','AutoVolume未记忆修改'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    if not d(resourceId="android:id/switch_widget").info["checked"]:
        print("AutoVolume关闭成功")
    else:
        print("AutoVolume关闭失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AutoVolume可修改，且记忆修改','错误','AutoVolume关闭失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AutoVolume可修改，且记忆修改','错误','AutoVolume关闭失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AutoVolume可修改，且记忆修改','正确','Sound_AutoVolume可修改，且记忆修改'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AutoVolume可修改，且记忆修改','正确','Sound_AutoVolume可修改，且记忆修改'])
    workbook.save(file_path_Sound) 
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_AutoVolume.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    for i in range(2):
        d.press("back")
        time.sleep(1)

##Speaker
def Sound_Speaker():
    os.system("start d:\logcat.bat Sound_Speaker")
    global sheet
    global workbook
    global file_path_Sound
    into_Sound()
    
    for i in range(6):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Auto":
        print("Sound_Speaker默认Auto")
    else:
        print("Sound_Speaker默认非Auto")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker默认非Auto'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker默认非Auto'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "On":
        print("Sound_Speaker修改成功")
    else:
        print("Sound_Speaker修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    
    into_Sound()
    for i in range(6):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "On":
        print("Sound_Speaker记忆修改")
    else:
        print("Sound_Speaker未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker未记忆修改'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Off":
        print("Sound_Speaker修改成功")
    else:
        print("Sound_Speaker修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    
    into_Sound()
    for i in range(6):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Off":
        print("Sound_Speaker记忆修改")
    else:
        print("Sound_Speaker未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker未记忆修改'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Auto":
        print("Sound_Speaker修改成功")
    else:
        print("Sound_Speaker修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Speaker可修改，且记忆修改','错误','Sound_Speaker修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Sound_Speaker可修改，且记忆修改','正确','Sound_Speaker可修改，且记忆修改'))
    sheet.append([test_num, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),'Sound_Speaker可修改，且记忆修改','正确','Sound_Speaker可修改，且记忆修改'])
    workbook.save(file_path_Sound) 
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_Speaker.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    for i in range(2):
        d.press("back")
        time.sleep(1)
        
##Audio Type
def Sound_AudioType():
    os.system("start d:\logcat.bat Sound_AudioType")
    global sheet
    global workbook
    global file_path_Sound
    into_Sound()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Normal":
        print("Sound_AudioType默认Normal")
    else:
        print("Sound_AudioType默认非Normal")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','Sound_AudioType默认非Normal'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','Sound_AudioType默认非Normal'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]').info["enabled"] == "false":
        print("AudioType的Normal模式下Visually Impaired不可选中")
    else:
        print("AudioType的Normal模式下Visually Impaired可选中")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','AudioType的Normal模式下Visually Impaired可选中'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','AudioType的Normal模式下Visually Impaired可选中'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Visually Impaired":
        print("Sound_AudioType修改成功")
    else:
        print("Sound_AudioType修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','Sound_AudioType修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','Sound_AudioType修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]').info["enabled"] == "true":
        print("AudioType的Visually Impaired模式下Visually Impaired可选中")
    else:
        print("AudioType的Visually Impaired模式下Visually Impaired不可选中")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','AudioType的Normal模式下Visually Impaired不可选中'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','AudioType的Normal模式下Visually Impaired不可选中'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
    
    into_Sound()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Visually Impaired":
        print("Sound_AudioType记忆修改")
    else:
        print("Sound_AudioType未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','Sound_AudioType记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','Sound_AudioType未记忆修改'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]').info["enabled"] == "true":
        print("AudioType的Visually Impaired模式下Visually Impaired可选中")
    else:
        print("AudioType的Normal模式下Visually Impaired不可选中")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','AudioType的Normal模式下Visually Impaired不可选中'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','AudioType的Normal模式下Visually Impaired不可选中'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("center")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Normal":
        print("Sound_AudioType修改成功")
    else:
        print("Sound_AudioType修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','Sound_AudioType修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','Sound_AudioType修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]').info["enabled"] == "false":
        print("AudioType的Normal模式下Visually Impaired不可选中")
    else:
        print("AudioType的Normal模式下Visually Impaired可选中")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','AudioType的Normal模式下Visually Impaired不可选中'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','错误','AudioType的Normal模式下Visually Impaired不可选中'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','正确','Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中','正确','Sound_AudioType可修改，且记忆修改,Normal状态下Visually Impaired不可选中'])
    workbook.save(file_path_Sound) 
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_AudioType.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    
##Visually Impaired
def Sound_VisuallyImpired():
    os.system("start d:\logcat.bat Sound_VisuallyImpaired")
    global sheet
    global workbook
    global file_path_Sound
    into_Sound()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Normal":
        d.press("center")
        time.sleep(1)
        d.press("down")
        time.sleep(1)
        d.press("center")
        time.sleep(1)
        if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Visually Impaired":
            print("Audio Type修改Visually Impaired成功")
        else:
            print("Audio Type修改Visually Impaired失败")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Audio Type修改Visually Impaired失败'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Audio Type修改Visually Impaired失败'])
            workbook.save(file_path_Sound) 
            os.system("adb shell kill `pgrep logcat`")
            return
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d(resourceId="android:id/switch_widget").info["checked"] and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/preference_progress_layout"]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "60":
        print("Sound_VisuallyImpaired默认参数正确")
    else:
        print("Sound_VisuallyImpaired默认参数错误")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Sound_VisuallyImpaired默认参数错误'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Sound_VisuallyImpaired默认参数错误'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]').info["enabled"] == "true":
        print("Visually Impaired_Speaker开启状态Volume可选中")
    else:
        print("Visually Impaired_Speaker开启状态Volume不可选中")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Visually Impaired_Speaker开启状态Volume可选中'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Visually Impaired_Speaker开启状态Volume可选中'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    if not d(resourceId="android:id/switch_widget").info["checked"]:
        print("Speaker关闭成功")
    else:
        print("Speaker关闭失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Speaker关闭失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Speaker关闭失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]').info["enabled"] == "false":
        print("Visually Impaired_Speaker关闭状态Volume不可选中")
    else:
        print("Visually Impaired_Speaker关闭状态Volume可选中")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Visually Impaired_Speaker关闭状态Volume可选中'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Visually Impaired_Speaker关闭状态Volume可选中'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    if d(resourceId="android:id/switch_widget").info["checked"]:
        print("Speaker开启成功")
    else:
        print("Speaker开启失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Speaker开启失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Speaker开启失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]').info["enabled"] == "true":
        print("Visually Impaired_Speaker开启状态Volume可选中")
    else:
        print("Visually Impaired_Speaker开启状态Volume不可选中")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Visually Impaired_Speaker开启状态Volume可选中'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Visually Impaired_Speaker开启状态Volume可选中'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    Volume = int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/preference_progress_layout"]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text())
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/preference_progress_layout"]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Volume - 5:
        print("Volume修改成功")
    else:
        print("Volume修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Volume修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Volume修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("back")
        time.sleep(1)
    
    into_Sound()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d(resourceId="android:id/switch_widget").info["checked"] and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/preference_progress_layout"]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Volume - 5:
        print("Sound_Visually Impared参数记忆修改")
    else:
        print("Sound_Visually Impared参数未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Sound_Visually Impared参数未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Sound_Visually Impared参数未记忆修改'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("down")
    time.sleep(1)
    for i in range(10):
        d.press("right")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/preference_progress_layout"]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Volume + 5:
        print("Volume修改成功")
    else:
        print("Volume修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Volume修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Volume修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(3):
        d.press("back")
        time.sleep(1)
        
    into_Sound()
    for i in range(8):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d(resourceId="android:id/switch_widget").info["checked"] and int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/preference_progress_layout"]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Volume + 5:
        print("Sound_Visually Impared参数记忆修改")
    else:
        print("Sound_Visually Impared参数未记忆修改")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Sound_Visually Impared参数未记忆修改'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Sound_Visually Impared参数未记忆修改'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("down")
    time.sleep(1)
    for i in range(5):
        d.press("left")
        time.sleep(1)
    if int(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/preference_progress_layout"]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text()) == Volume:
        print("Volume修改成功")
    else:
        print("Volume修改失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Volume修改失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Volume修改失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("up")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if not d(resourceId="android:id/switch_widget").info["checked"]:
        print("Sound_Visually Impaired_Speaker关闭成功")
    else:
        print("Sound_Visually Impaired_Speaker关闭失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Sound_Visually Impaired_Speaker关闭失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','错误','Sound_Visually Impaired_Speaker关闭失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','正确','Sound_Visually Impaired可修改，且记忆修改'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Visually Impaired可修改，且记忆修改','正确','Sound_Visually Impaired可修改，且记忆修改'])
    workbook.save(file_path_Sound) 
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_VisuallyImpaired.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    d.press("up")
    time.sleep(1)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    for i in range(2):
        d.press("back")
        time.sleep(1)
        
##Dialogue Enhancer
def Sound_DialogueEnhancer():
    os.system("start d:\logcat.bat Sound_DialogueEnhancer")
    global sheet
    global workbook
    global file_path_Sound
    into_Sound()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Normal":
        d.press("down")
        time.sleep(1)
    else:
        for i in range(2):
            d.press("down")
            time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Off":
        print("Sound_DialogueEnhancer默认Off")
    else:
        print("Sound_DialogueEnhancer默认非Off")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer默认非Off'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer默认非Off'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("down")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Low":
        print("Sound_DialogueEnhancer修改Low成功")
    else:
        print("Sound_DialogueEnhancer修改Low失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer修改Low失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer修改Low失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
        
    into_Sound()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Normal":
        d.press("down")
        time.sleep(1)
    else:
        for i in range(2):
            d.press("down")
            time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Low":
        print("Sound_DialogueEnhancer记忆修改Low")
    else:
        print("Sound_DialogueEnhancer未记忆修改Low")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer未记忆修改Low'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer未记忆修改Low'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    for i in range(2):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Medium":
        print("Sound_DialogueEnhancer修改Medium成功")
    else:
        print("Sound_DialogueEnhancer修改Medium失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer修改Medium失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer修改Medium失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
        
    into_Sound()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Normal":
        d.press("down")
        time.sleep(1)
    else:
        for i in range(2):
            d.press("down")
            time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Medium":
        print("Sound_DialogueEnhancer记忆修改Medium")
    else:
        print("Sound_DialogueEnhancer未记忆修改Medium")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer未记忆修改Medium'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer未记忆修改Medium'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    for i in range(3):
        d.press("down")
        time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "High":
        print("Sound_DialogueEnhancer修改High成功")
    else:
        print("Sound_DialogueEnhancer修改High失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer修改High失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer修改High失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    for i in range(2):
        d.press("back")
        time.sleep(1)
        
    into_Sound()
    for i in range(7):
        d.press("down")
        time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Normal":
        d.press("down")
        time.sleep(1)
    else:
        for i in range(2):
            d.press("down")
            time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "High":
        print("Sound_DialogueEnhancer记忆修改High")
    else:
        print("Sound_DialogueEnhancer未记忆修改High")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer未记忆修改High'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer未记忆修改High'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("center")
    time.sleep(1)
    d.press("center")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Off":
        print("Sound_DialogueEnhancer修改Off成功")
    else:
        print("Sound_DialogueEnhancer修改Off失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','错误','Sound_DialogueEnhancer修改Off失败'))
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','正确','Sound_DialogueEnhancer可修改，且记忆修改'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_DialogueEnhancer可修改，且记忆修改','正确','Sound_DialogueEnhancer可修改，且记忆修改'])
    workbook.save(file_path_Sound) 
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_DialogueEnhancer.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    for i in range(3):
        d.press("back")
        time.sleep(1)

##Sound Reset
def Sound_Reset():
    os.system("start d:\logcat.bat Sound_Reset")
    global sheet
    global workbook
    global file_path_Sound
    into_Sound()
    d.press("up")
    time.sleep(1)
    for i in range(2):
        d.press("center")
        time.sleep(1)
    time.sleep(10)
    d.press("down")
    time.sleep(1)
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]').info["enabled"] == "true" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[1]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Standard":
        print("Sound Mode重置成功")
    else:
        print("Sound Mode重置失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Sound Mode重置失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Sound Mode重置失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]').info["enabled"] == "true" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "0":
        print("Balance重置成功")
    else:
        print("Balance重置失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Balance重置失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Balance重置失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]').info["enabled"] == "true" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50":
        print("Bass重置成功")
    else:
        print("Bass重置失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Bass重置失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Bass重置失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]').info["enabled"] == "true" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.RelativeLayout[1]/android.widget.TextView[1]').get_text() == "50":
        print("Treble重置成功")
    else:
        print("Treble重置失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Treble重置失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Treble重置失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]').info["enabled"] == "true":
        for i in range(4):
            d.press("down")
            time.sleep(1)
        d.press("center")
        time.sleep(1)
        if not (d(resourceId="android:id/switch_widget")[0].info["checked"] or d(resourceId="android:id/switch_widget")[1].info["checked"] or d(resourceId="android:id/switch_widget")[2].info["checked"]):
            print("DTS Studio Sound重置成功")
        else:
            print("DTS Studio Sound重置失败")
            table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','DTS Studio Sound重置失败'))
            sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','DTS Studio Sound重置失败'])
            workbook.save(file_path_Sound) 
            os.system("adb shell kill `pgrep logcat`")
            return
    else:
        print("DTS Studio Sound重置失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','DTS Studio Sound重置失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','DTS Studio Sound重置失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    d.press("back")
    time.sleep(1)
    
    for i in range(4):
        d.press("down")
        time.sleep(1)
    
    if not d(resourceId="android:id/switch_widget").info["checked"]:
        print("Auto Volume重置成功")
    else:
        print("Auto Volume重置失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Auto Volume重置失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Auto Volume重置失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]').info["enabled"] == "true" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Auto":
        print("Speaker重置成功")
    else:
        print("Speaker重置失败")
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]').info["enabled"])
        print(d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[3]/android.widget.LinearLayout[1]/android.widget.TextView[1]').get_text())
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Speaker重置失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Speaker重置失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]').info["enabled"] == "true" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[4]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Normal":
        print("Audio Type重置成功")
    else:
        print("Audio Type重置失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Audio Type重置失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Audio Type重置失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[5]').info["enabled"] == "false":
        print("Visually Impaired重置成功")
    else:
        print("Visually Impaired重置失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Visually Impaired重置失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Visually Impaired重置失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    if d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]').info["enabled"] == "true" and d.xpath('//*[@resource-id="com.smartdevice.livetv:id/list"]/android.widget.LinearLayout[6]/android.widget.LinearLayout[1]/android.widget.TextView[2]').get_text() == "Off":
        print("Dialogue Enhancer重置成功")
    else:
        print("Dialogue Enhancer重置失败")
        table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Dialogue Enhancer重置失败'))
        sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','错误','Dialogue Enhancer重置失败'])
        workbook.save(file_path_Sound) 
        os.system("adb shell kill `pgrep logcat`")
        return
    table.insert('','end',value = (test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','正确','Sound_Reset重置所有数据'))
    sheet.append([test_num,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), 'Sound_Reset重置所有数据','正确','Sound_Reset重置所有数据'])
    workbook.save(file_path_Sound) 
    os.system("adb shell kill `pgrep logcat`")
    try:
        os.remove('D:\\auto_test_log\Sound_Reset.log')
        print('文件删除成功！')
    except Exception as e:
        print('文件删除失败:', e)
    for i in range(3):
        d.press("back")
        time.sleep(1)
    
##用于测试
def test():
    
    d.press("home")
    time.sleep(1)
    d.shell("am start -n com.smartdevice.livetv/skyworth.skyworthlivetv.osd.ui.mainActivity.LiveTvScreenActivity")
    time.sleep(5)
    d.shell("input keyevent 13")
    time.sleep(5)
    d.shell("am start -n com.smartdevice.livetv/com.product.osd.AndroidTvSettings.PictureActivity")
    time.sleep(5)
    d.press("center")
    time.sleep(1)
    print((d(className = "android.widget.LinearLayout").child(className = "android.widget.RelativeLayout")).info["focused"])
    print((d(className = "android.widget.LinearLayout").child(text = "Picture Mode")).info["focused"])
    
    #d(className="android.widget.LinearLayout").child(className="android.widget.RelativeLayout").child(text = "Picture Mode").long_click()
    while not bool(d(className="android.widget.LinearLayout").child(className="android.widget.LinearLayout").child(text = "Gamma")):
        d.press("down")
        time.sleep(1)

test_num = "选择测试用例"

def Func_auto_or_self(event):
    print(auto_or_self.get())
    if auto_or_self.get() == "手动测试":
        choose_test['value'] = temp_list_All
        choose_test.state(['!disabled'])
    elif auto_or_self.get() == "自动测试所有":
        choose_test['value'] = temp_list_All
        choose_test.state(['disabled'])
    elif auto_or_self.get() == "自动测试Picture":
        choose_test['value'] = temp_list_Picture
        choose_test.state(['disabled'])
    elif auto_or_self.get() == "自动测试Sound":
        choose_test['value'] = temp_list_Sound
        choose_test.state(['disabled'])
        
def Func_choose_test(event):
    print(choose_test.get())
    global test_num
    test_num = str(choose_test.get())
        
def Func_start_test():
    global test_num
    if auto_or_self.get() == "自动测试全部":
        test_num = "Picture_PictureMode"
        Picture_PictureMode()
        test_num = "into_Picture_Mode"
        into_Picture_Mode()
        test_num = "Picture_ScreenBrightness"
        Picture_ScreenBrightness()
        test_num = "Picture_Brightness"
        Picture_Brightness()
        test_num = "Picture_Contrast"
        Picture_Contrast()
        test_num = "Picture_Saturation"
        Picture_Saturation()
        test_num = "Picture_HUE"
        Picture_HUE()
        test_num = "Picture_Sharpness"
        Picture_Sharpness()
        test_num = "Picture_AutoBrightness"
        Picture_AutoBrightness()
        test_num = "Picture_Gamma"
        Picture_Gamma()
        test_num = "Picture_ColorTemperature_Mode"
        Picture_ColorTemperature()
        test_num = "Picture_ColorTemperature_modify"
        Picture_ColorTemperature_User()
        test_num = "Picture_ColorTemperature_modify_to_user"
        Picture_ColorTemperature_()
        test_num = "Picture_PWBC"
        Picture_PWBC()
        test_num = "Picture_AdvancedSettings"
        Picture_AdvancedSettings()
        test_num = "Picture_ColorTuner"
        Picture_ColorTuner()
        test_num = "Picture_Reset"
        Picture_Reset()
        test_num = "Picture_Mode_compare_screen"
        Picture_Mode_compare_screen()
        test_num = "Picture_Mode_cool_Temperature_compare_screen"
        Picture_Mode_cool_Temperature_compare_screen()
        
        test_num = "Sound_SoundMode"
        Sound_SoundMode()
        test_num = "Sound_Balance_Bass_Treble"
        Sound_Balance_Bass_Treble()
        test_num = "Sound_DTS"
        Sound_DTS()
        test_num = "Sound_AutoVolume"
        Sound_AutoVolume()
        test_num = "Sound_Speaker"
        Sound_Speaker()
        test_num = "Sound_AudioType"
        Sound_AudioType()
        test_num = "Sound_VisuallyImpired"
        Sound_VisuallyImpired()
        test_num = "Sound_DialogueEnhancer"
        Sound_DialogueEnhancer()
        test_num = "Sound_Reset"
        Sound_Reset()
    elif auto_or_self.get() == "手动测试":
        if test_num == "Picture_PictureMode":
            Picture_PictureMode()
        elif test_num == "into_Picture_Mode":
            into_Picture_Mode()
        elif test_num == "Picture_ScreenBrightness":
            Picture_ScreenBrightness()
        elif test_num == "Picture_Brightness":
            Picture_Brightness()
        elif test_num == "Picture_Contrast":
            Picture_Contrast()
        elif test_num == "Picture_Saturation":
            Picture_Saturation()
        elif test_num == "Picture_HUE":
            Picture_HUE()
        elif test_num == "Picture_Sharpness":
            Picture_Sharpness()
        elif test_num == "Picture_AutoBrightness":
            Picture_AutoBrightness()
        elif test_num == "Picture_Gamma":
            Picture_Gamma()
        elif test_num == "Picture_ColorTemperature_Mode":
            Picture_ColorTemperature()
        elif test_num == "Picture_ColorTemperature_modify":
            Picture_ColorTemperature_User()
        elif test_num == "Picture_ColorTemperature_modify_to_user":
            Picture_ColorTemperature_()
        elif test_num == "Picture_PWBC":
            Picture_PWBC()
        elif test_num == "Picture_AdvancedSettings":
            Picture_AdvancedSettings()
        elif test_num == "Picture_ColorTuner":
            Picture_ColorTuner()
        elif test_num == "Picture_Reset":
            Picture_Reset()
        elif test_num == "Picture_Mode_compare_screen":
            Picture_Mode_compare_screen()
        elif test_num == "Picture_Mode_cool_Temperature_compare_screen":
            Picture_Mode_cool_Temperature_compare_screen()
            
        if test_num == "Sound_Default":
            Sound_Default()
        elif test_num == "Sound_DolbyAtmos":
            Sound_DolbyAtmos()
        elif test_num == "Sound_SoundMode":
            Sound_SoundMode()
        elif test_num == "Sound_DolbySoundMode":
            Sound_DolbySoundMode()
        elif test_num == "Sound_Balance_Bass_Treble":
            Sound_Balance_Bass_Treble()
        elif test_num == "Sound_DTS":
            Sound_DTS()
        elif test_num == "Sound_AutoVolume":
            Sound_AutoVolume()
        elif test_num == "Sound_Speaker":
            Sound_Speaker()
        elif test_num == "Sound_AudioType":
            Sound_AudioType()
        elif test_num == "Sound_VisuallyImpired":
            Sound_VisuallyImpired()
        elif test_num == "Sound_DialogueEnhancer":
            Sound_DialogueEnhancer()
        elif test_num == "Sound_Reset":
            Sound_Reset()
        elif test_num == "test":
            test()
    elif auto_or_self.get() == "自动测试Picture":
        if test_num == "Picture_PictureMode":
            Picture_PictureMode()
        elif test_num == "into_Picture_Mode":
            into_Picture_Mode()
        elif test_num == "Picture_ScreenBrightness":
            Picture_ScreenBrightness()
        elif test_num == "Picture_Brightness":
            Picture_Brightness()
        elif test_num == "Picture_Contrast":
            Picture_Contrast()
        elif test_num == "Picture_Saturation":
            Picture_Saturation()
        elif test_num == "Picture_HUE":
            Picture_HUE()
        elif test_num == "Picture_Sharpness":
            Picture_Sharpness()
        elif test_num == "Picture_AutoBrightness":
            Picture_AutoBrightness()
        elif test_num == "Picture_Gamma":
            Picture_Gamma()
        elif test_num == "Picture_ColorTemperature_Mode":
            Picture_ColorTemperature()
        elif test_num == "Picture_ColorTemperature_modify":
            Picture_ColorTemperature_User()
        elif test_num == "Picture_ColorTemperature_modify_to_user":
            Picture_ColorTemperature_()
        elif test_num == "Picture_PWBC":
            Picture_PWBC()
        elif test_num == "Picture_AdvancedSettings":
            Picture_AdvancedSettings()
        elif test_num == "Picture_ColorTuner":
            Picture_ColorTuner()
        elif test_num == "Picture_Reset":
            Picture_Reset()
        elif test_num == "Picture_Mode_compare_screen":
            Picture_Mode_compare_screen()
        elif test_num == "Picture_Mode_cool_Temperature_compare_screen":
            Picture_Mode_cool_Temperature_compare_screen()
        elif test_num == "test":
            test()
    elif auto_or_self.get() == "自动测试Sound":
        if test_num == "Sound_Default":
            Sound_Default()
        elif test_num == "Sound_DolbyAtmos":
            Sound_DolbyAtmos()
        elif test_num == "Sound_SoundMode":
            Sound_SoundMode()
        elif test_num == "Sound_DolbySoundMode":
            Sound_DolbySoundMode()
        elif test_num == "Sound_Balance_Bass_Treble":
            Sound_Balance_Bass_Treble()
        elif test_num == "Sound_DTS":
            Sound_DTS()
        elif test_num == "Sound_AutoVolume":
            Sound_AutoVolume()
        elif test_num == "Sound_Speaker":
            Sound_Speaker()
        elif test_num == "Sound_AudioType":
            Sound_AudioType()
        elif test_num == "Sound_VisuallyImpired":
            Sound_VisuallyImpired()
        elif test_num == "Sound_DialogueEnhancer":
            Sound_DialogueEnhancer()
        elif test_num == "Sound_Reset":
            Sound_Reset()
        elif test_num == "test":
            test()
            
def Func_end_test():
    top.destroy()
    d.press("back")
    time.sleep(1)
    d.press("back")
    time.sleep(1)
    d.press("back")
    time.press(2)
    d.press("home")
    time.press(2)
    
auto_or_self.bind("<<ComboboxSelected>>",Func_auto_or_self)

choose_test.bind("<<ComboboxSelected>>",Func_choose_test)

stop_flag = True

def thread_it(func, *args):
    t = threading.Thread(target = func, args = args)
    t.setDaemon(True)
    t.start()

start_test = ttk.Button(frm_right, text = "开始测试", command = lambda :thread_it(Func_start_test))
start_test.pack()

end_test = ttk.Button(frm_right, text = "退出", command = Func_end_test)
end_test.pack()

frm.pack()
frm_left.pack(side = LEFT)
frm_right.pack(side = RIGHT)

table = ttk.Treeview(
    master = top,
    height = 25,
    columns = ['测试用例编号','测试时间','预期结果','测试结果','原因'],
    show = 'headings'
    )
table.heading('测试用例编号',text = '测试用例')
table.heading('测试时间',text = '测试时间')
table.heading('预期结果',text = '预期结果')
table.heading('测试结果',text = '测试结果')
table.heading('原因',text = '原因')

table.column('测试用例编号',width = 180,minwidth = 180,anchor = 'center')
table.column('测试时间',width = 80,minwidth = 80,anchor = 'center')
table.column('预期结果',width = 480,minwidth = 480,anchor = 'center')
table.column('测试结果',width = 80,minwidth = 80,anchor = 'center')
table.column('原因',width = w - 800,minwidth = w - 800,anchor = 'center')
    
note = ttk.Label(top,text = "注：检测数据存放于D:\Picture.xlsx,出错日志存放于D:\\auto_test_log下")
note.pack()
    
table.pack(side = TOP)



top.mainloop()

## -transparentcolor, -disabled, -fullscreen, -toolwindow, or -topmost

#d.xpath('//*[@resource-id="com.smartdevice.livetv:id/btn_scan_channel"]/android.widget.Button').longClick()

"""
from openpyxl import load_workbook
from openpyxl import Workbook
import os
# 指定文件路径
file_path = "D:/Picture.xlsx"

try:
    # 尝试打开现有的工作簿
    workbook = load_workbook(file_path)
    sheet = workbook.active
except FileNotFoundError:
    # 如果文件不存在，则创建一个新的工作簿
    workbook = Workbook()
    sheet = workbook.active
    data = ["测试用例编号", "测试时间", "预期结果","测试结果","原因"]
    sheet.append(data)
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["A"].width = 50
    sheet.column_dimensions["A"].width = 100
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["A"].width = 100

# 在最后一行添加数据
for i in range(2):
    data = ["Data 1", "Data 2", "Data 3", "Data 4", "Data 5"]
    sheet.append(data)

data = ["Data 1", "Data 2", "Data 3", "Data 4", "Data 5"]
sheet.append(data)
# 保存工作簿
workbook.save(file_path)
"""


